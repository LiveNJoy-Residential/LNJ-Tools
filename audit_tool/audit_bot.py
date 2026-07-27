"""
ResMan Concession & Recurring Revenue Integrity Audit Bot
==========================================================
Authors  : Concession Audit Engine + Revenue Integrity Engine
Company  : LiveNjoy Residential  |  System: ResMan  |  Period: February 2026

FILE â†’ DATA FOLDER MAPPING
---------------------------
data/transactions/  â† Transaction List Reports (Credits)
                       e.g. "CAI Transaction List (Credits) - Feb 2026.csv"
                       â†’ Concession Audit Engine: actual concession postings

data/leases/        â† New and Renewed Leases
                       e.g. "Crossing at Irving New and Renewed Leases.csv"
                       â†’ Concession Audit Engine: legal approved concession amounts

data/edits/         â† Edited Transactions by User
                       e.g. "Crossing at Irving Edited Transactions by User.csv"
                       â†’ Override Audit: who reversed / changed what

data/recurring/     â† Transaction Projections (Recurring)
                       e.g. "Crossings at Irving Recurring Transaction Projection.csv"
                       â†’ Daniel Stage 1: what SHOULD post each month per unit

data/rent_rolls/    â† Rent Rolls
                       e.g. "Crossings at Irving Rent Roll.csv"
                       â†’ Daniel Stage 2: what IS actually configured per unit

data/activity/      â† Resident Activity reports
                       e.g. "Crossing at Irving Resident Activity.csv"
                       â†’ Move-in dates, current lease dates, assigned manager

NOTE: Resident Ledgers are PDFs -- cannot be auto-processed.
      They live in exports/Resident Ledgers/ for manual review only.
"""

import io
import os
import re
import warnings
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------------
# GLOBAL PATHS
# ---------------------------------------------------------------------------
BASE_DIR   = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR   = os.path.join(BASE_DIR, "data")
OUTPUT_DIR = os.path.join(BASE_DIR, "output")

DIRS = {
    "transactions":       os.path.join(DATA_DIR, "transactions"),         # Transaction List (Credits)
    "leases":             os.path.join(DATA_DIR, "leases"),               # New & Renewed Leases
    "edits":              os.path.join(DATA_DIR, "edits"),                # Edited Transactions by User
    "recurring":          os.path.join(DATA_DIR, "recurring"),            # Transaction Projections
    "rent_rolls":         os.path.join(DATA_DIR, "rent_rolls"),           # Rent Roll
    "activity":           os.path.join(DATA_DIR, "activity"),             # Resident Activity
    "market_rent_schedule": os.path.join(DATA_DIR, "market rent schedule"), # Market Rent Schedule Detail
}

# ---------------------------------------------------------------------------
# AUDIT CONSTANTS
# ---------------------------------------------------------------------------
APPROVED_CODES = {"CONR", "CRTCO", "EMPL", "MCCR", "RRFee"}

# How approved concessions appear in ResMan Description field
APPROVED_DESCRIPTIONS = {
    "concession - rent",
    "courtesy officer",
    "employee unit rent allowance",
    "miscellaneous credit",
    "resident referral",
}

CONCESSION_CRITICAL_AMT   = 700
CONCESSION_HIGH_AMT       = 500
CONCESSION_HIGH_MONTHS    = 2
STANDARD_CHARGE_THRESHOLD = 0.90
RECENT_MOVEIN_DAYS        = 60

AUDIT_MONTH = "Jun 2026"   # update each month

# Categories exempt from "Missing Standard Charge" â€” these are unit-specific add-ons.
# Per Daniel Twito: parking, pet fees, and washer/dryer are not universal.
OPTIONAL_CHARGE_KEYWORDS = {
    "carport", "parking", "pet rent", "pet fee", "washer", "dryer",
    "first floor", "1st floor",
}

# Net Effective Rent floors by property and bedroom type (Daniel Twito, June 2026).
# NER = recurring_rent + recurring_concession  (concession is stored negative).
# Flag when NER < floor â€” catches double-discounts and below-floor net rents.
PROPERTY_NER_FLOORS = {
    "Parks on Taylor":     {"1BR": 799, "2BR": 849},   # 2BR floor updated 4/30/2026
    "Highland Park":       {"1BR": 799, "2BR": 999},
    "Valencia Plaza":      {"1BR": 999, "2BR": 999},
    "Western Station":     {"1BR": None, "2BR": None},   # floor varies by floorplan
    "Village Green":       {"1BR": None, "2BR": None},   # $300 off market â€” see PROPERTY_NER_DISCOUNT
    "Crossings at Irving": {"1BR": None, "2BR": None},   # $100 off market â€” see PROPERTY_NER_DISCOUNT
    "La Prada":            {"1BR": None, "2BR": None},   # $100 off market â€” see PROPERTY_NER_DISCOUNT
}

# For properties whose NER floor is dynamic (market rent minus a discount),
# this dict maps property -> max allowed discount from market rent.
# Floor = per-unit market rent (from Market Rent Schedule) - discount.
# NER below that threshold is flagged as Net Effective Rent Below Floor.
PROPERTY_NER_DISCOUNT = {
    "Village Green":       300,   # $300 off market (per Daniel Twito, CONTEXT.md)
    "Crossings at Irving": 100,   # $100 off market
    "La Prada":            100,   # $100 off market
    # Western Station: floor varies by floorplan â€” not a simple market-minus rule; skip
}

# â”€â”€â”€ Official monthly fee schedule per property â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
# Source: Fee sheet .docx files provided by Daniel Twito (March 2026).
# Used to validate charge AMOUNTS in the Recurring Projection.
# Per Daniel: $1 is the variance cutoff for charge amount comparisons.
# NOTE: La Prada omitted â€” no fee sheet provided.
# â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
PROPERTY_FEE_SCHEDULE = {
    "Crossings at Irving": [
        {"name": "Billing Fee",         "keywords": ["billing fee"],                       "amount": 5.00,  "optional": False},
        {"name": "Trash Service",        "keywords": ["trash service"],                     "amount": 10.00, "optional": False},
        {"name": "Pest Control",         "keywords": ["pest control"],                      "amount": 8.00,  "optional": False},
        {"name": "Package Locker",       "keywords": ["package locker"],                    "amount": 9.00,  "optional": False},
        {"name": "Cable/Internet (Spectrum)",  "keywords": ["cable", "internet", "spectrum"],  "amount": 55.00, "optional": False},
        {"name": "First Floor Fee",      "keywords": ["first floor", "1st floor"],          "amount": 25.00, "optional": True},
        {"name": "Washer/Dryer Rental",  "keywords": ["washer", "dryer"],                   "amount": 55.00, "optional": True},
        {"name": "Reserved Parking",     "keywords": ["carport", "parking"],                "amount": 35.00, "optional": True},
        {"name": "Pet Rent",             "keywords": ["pet rent"],                          "amount": 25.00, "optional": True},
    ],
    "Highland Park": [
        {"name": "Billing Fee",         "keywords": ["billing fee"],                       "amount": 5.00,  "optional": False},
        {"name": "Trash Service",        "keywords": ["trash service"],                     "amount": 15.00, "optional": False},
        {"name": "Pest Control",         "keywords": ["pest control"],                      "amount": 5.00,  "optional": False},
        {"name": "Community Fee (CAM)",  "keywords": ["cam", "community fee"],              "amount": 10.00, "optional": True},
        {"name": "Valet Trash",          "keywords": ["valet trash"],                       "amount": 35.00, "optional": True},
        {"name": "Washer/Dryer Rental",  "keywords": ["washer", "dryer"],                   "amount": 50.00, "optional": True},
        {"name": "Reserved Parking",     "keywords": ["carport", "parking"],                "amount": 35.00, "optional": True},
        {"name": "Pet Rent",             "keywords": ["pet rent"],                          "amount": 15.00, "optional": True},
    ],
    "Parks on Taylor": [
        {"name": "Billing Fee",         "keywords": ["billing fee"],                       "amount": 5.00,  "optional": False},
        {"name": "Trash Service",        "keywords": ["trash service"],                     "amount": 15.00, "optional": False},
        {"name": "Pest Control",         "keywords": ["pest control"],                      "amount": 5.00,  "optional": False},
        {"name": "Community Fee (CAM)",  "keywords": ["cam", "community fee"],              "amount": 10.00, "optional": False},
        {"name": "Washer/Dryer Rental",  "keywords": ["washer", "dryer"],                   "amount": 50.00, "optional": True},
        {"name": "Reserved Parking",     "keywords": ["carport", "parking"],                "amount": 25.00, "optional": True},
        {"name": "Pet Rent",             "keywords": ["pet rent"],                          "amount": 20.00, "optional": True},
    ],
    "Valencia Plaza": [
        {"name": "Billing Fee",         "keywords": ["billing fee", "utility admin fee"],   "amount": 5.00,  "optional": False},
        {"name": "Trash Service",        "keywords": ["trash service"],                     "amount": 10.00, "optional": False},
        {"name": "Pest Control",         "keywords": ["pest control"],                      "amount": 6.00,  "optional": False},
        {"name": "Washer/Dryer Rental",  "keywords": ["washer", "dryer"],                   "amount": 50.00, "optional": True},
        {"name": "Reserved Parking",     "keywords": ["carport", "parking"],                "amount": 35.00, "optional": True},
        {"name": "Pet Rent",             "keywords": ["pet rent"],                          "amount": 20.00, "optional": True},
    ],
    "Village Green": [
        {"name": "Billing Fee",         "keywords": ["billing fee"],                       "amount": 5.00,  "optional": False},
        {"name": "Trash Service",        "keywords": ["trash service"],                     "amount": 10.00, "optional": False},
        {"name": "Pest Control",         "keywords": ["pest control"],                      "amount": 8.00,  "optional": False},
        {"name": "Community Fee (CAM)",  "keywords": ["cam", "community fee"],              "amount": 10.00, "optional": False},
        {"name": "Valet Trash",          "keywords": ["valet trash"],                       "amount": 35.00, "optional": False},
        {"name": "HOA Fee",              "keywords": ["hoa"],                               "amount": 2.50,  "optional": False},
        {"name": "Package Locker",       "keywords": ["package locker"],                    "amount": 9.00,  "optional": False},
        {"name": "Cable/Internet",       "keywords": ["cable", "internet"],                 "amount": 55.00, "optional": False},
        {"name": "Washer/Dryer Rental",  "keywords": ["washer", "dryer"],                   "amount": 50.00, "optional": True},
        {"name": "Carport Rental",       "keywords": ["carport", "parking"],                "amount": 35.00, "optional": True},
        {"name": "Pet Rent",             "keywords": ["pet rent"],                          "amount": 25.00, "optional": True},
    ],
    "Western Station": [
        {"name": "Billing Fee",         "keywords": ["billing fee"],                       "amount": 5.00,  "optional": False},
        {"name": "Trash Service",        "keywords": ["trash service"],                     "amount": 10.00, "optional": False},
        {"name": "Pest Control",         "keywords": ["pest control"],                      "amount": 10.00, "optional": False},
        {"name": "Community Fee (CAM)",  "keywords": ["cam", "community fee"],              "amount": 10.00, "optional": False},
        {"name": "Valet Trash",          "keywords": ["valet trash"],                       "amount": 35.00, "optional": False},
        {"name": "Package Locker",       "keywords": ["package locker"],                    "amount": 9.00,  "optional": False},
        {"name": "Washer/Dryer Rental",  "keywords": ["washer", "dryer"],                   "amount": 50.00, "optional": True},
        {"name": "Reserved Parking",     "keywords": ["carport", "parking"],                "amount": 50.00, "optional": True},
        {"name": "Pet Rent",             "keywords": ["pet rent"],                          "amount": 25.00, "optional": True},
    ],
    "La Prada": [
        {"name": "Billing Fee",          "keywords": ["billing fee"],                    "amount": 5.00,  "optional": False},
        {"name": "Trash Service",         "keywords": ["trash service"],                  "amount": 10.00, "optional": False},
        {"name": "Package Locker",        "keywords": ["package locker"],                 "amount": 7.50,  "optional": False},
        {"name": "Pest Control",          "keywords": ["pest control"],                   "amount": 6.00,  "optional": False},
        {"name": "Washer/Dryer Rental",   "keywords": ["washer", "dryer"],                "amount": 50.00, "optional": True},
        {"name": "Reserved Parking",      "keywords": ["carport", "parking"],             "amount": 35.00, "optional": True},
        {"name": "Pet Rent",              "keywords": ["pet rent"],                       "amount": 20.00, "optional": True},
    ],
}

RISK_CRITICAL = "CRITICAL"
RISK_HIGH     = "HIGH"
RISK_MEDIUM   = "MEDIUM"
RISK_VERIFY   = "VERIFY"

RISK_MAP = {
    # John's Rules
    "Post-Term Credit":               RISK_CRITICAL,
    "Recurring Concession >$700":     RISK_HIGH,    # single month â‰¥$700 = HIGH per Daniel (Jun 4); multi-month = CRITICAL via Full Rent Recurring Offset
    "Missing Addendum":               RISK_HIGH,       # credit posted, no RR setup â€” HIGH not CRITICAL per Daniel June 4
    "Missing Lease":                  RISK_HIGH,
    # "Invalid Credit Code" removed â€” John confirmed (March 10, 2026) that
    # descriptions are freeform identifiers with no standard rule; R4 disabled.
    "Concession Amount Mismatch":     RISK_HIGH,        # RR amount != TX amount
    "Not Properly Posted":            RISK_HIGH,        # RR setup but not in TX
    # Daniel's Stage 1
    "Missing Standard Charge":        RISK_HIGH,
    "Major Charge Amount Variance":   RISK_HIGH,
    "Concession >$500 for 2+ Months": RISK_HIGH,
    "Minor Charge Amount Variance":   RISK_MEDIUM,
    "Concession No Expiration":       RISK_MEDIUM,
    # Daniel's Stage 2
    "Negative Net Rent":              RISK_CRITICAL,
    "$0 Net Rent (Not Recent)":       RISK_CRITICAL,
    "Manual Posting Without Setup":   RISK_HIGH,
    "Posted vs Recurring Mismatch":   RISK_HIGH,
    "Misc Tenant Credit":             RISK_HIGH,
    "$0 Net Rent (Recent Move-in)":   RISK_MEDIUM,
    # Fee Schedule Check
    "Fee Schedule Violation":         RISK_HIGH,
    # NER Engine (Stage 1B) â€” Daniel Twito June 2026 methodology
    "Full Rent Recurring Offset":     RISK_CRITICAL,   # concession = rent for 2+ months
    "Net Effective Rent Below Floor": RISK_CRITICAL,   # NER below property floor
    "Double-Discount Setup":          RISK_CRITICAL,   # rent reduced + concession on top
    "Future Month Full Offset":       RISK_VERIFY,     # single-month free — verify addendum
    "Exact $500 Concession":          RISK_HIGH,       # unusual amount — may be first/last month split
    "NTV Proration":                  RISK_VERIFY,     # prorated rent for NTV unit
    "Recurring Rent Decrease":        RISK_HIGH,       # material drop not tied to NTV/move-out
    "Recurring Rent Decrease (NTV)":  RISK_VERIFY,     # rent drop likely due to NTV/move-out proration
}


# ===========================================================================
# SECTION 1 -- CLEANING HELPERS
# ===========================================================================

def clean_currency(val) -> float:
    """Strip $, commas, spaces -> float. Returns 0.0 on failure."""
    if pd.isna(val) or str(val).strip() in ("", "nan", "--"):
        return 0.0
    cleaned = re.sub(r'[$,"\s]', "", str(val))
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def clean_unit(val) -> str:
    """Normalize unit number: leading zeros stripped, '101 - Name' handled."""
    s = str(val).strip() if not pd.isna(val) else ""
    if not s or s.lower() == "nan":
        return "UNKNOWN"
    if " - " in s:
        s = s.split(" - ")[0].strip()
    return s.lstrip("0") or "0"


def parse_date(val):
    """Best-effort date parser; returns pd.Timestamp or None."""
    if pd.isna(val) or str(val).strip() in ("", "nan"):
        return None
    try:
        return pd.to_datetime(val, infer_datetime_format=True)
    except Exception:
        return None


def is_date_string(s: str) -> bool:
    """Return True if string looks like M/D/YYYY."""
    return bool(re.match(r"^\d{1,2}/\d{1,2}/\d{2,4}$", str(s).strip()))


def derive_property(filename: str) -> str:
    """
    Map any ResMan export filename to the standard full property name.
    Handles both short-code filenames (CAI, HP, ...) and long-name filenames.
    """
    CODE_MAP = {
        "CAI":  "Crossings at Irving",
        "POT":  "Parks on Taylor",
        "HP":   "Highland Park",
        "LP":   "La Prada",
        "VG":   "Village Green",
        "VPA":  "Valencia Plaza",
        "VP":   "Valencia Plaza",
        "WST":  "Western Station",
    }
    KEYWORD_MAP = {
        "crossing":  "Crossings at Irving",
        "irving":    "Crossings at Irving",
        "taylor":    "Parks on Taylor",
        "highland":  "Highland Park",
        "prada":     "La Prada",
        "village":   "Village Green",
        "valencia":  "Valencia Plaza",
        "western":   "Western Station",
    }
    first_word = filename.split(" ")[0].replace(",", "").upper()
    if first_word in CODE_MAP:
        return CODE_MAP[first_word]
    fname_lower = filename.lower()
    for keyword, prop in KEYWORD_MAP.items():
        if keyword in fname_lower:
            return prop
    return filename.split(" ")[0]  # fallback


def get_bedroom_type(unit_type: str) -> str:
    """Map ResMan unit type code to bedroom count label (1BR / 2BR / 3BR).
    Logic: if the numeric suffix is 3 or higher, classify as 3BR.
    Otherwise use the letter prefix (A = 1BR, B = 2BR, C/D = 3BR).
    Examples: A1 â†’1BR, A2 â†’1BR, B2 â†’2BR, B3 â†’3BR, B3-P â†’3BR, A1-P â†’1BR.
    """
    clean = unit_type.strip().upper()
    m = re.search(r'(\d+)', clean)
    if m and int(m.group(1)) >= 3:
        return "3BR"
    prefix = clean[0] if clean else ""
    return {"A": "1BR", "B": "2BR", "C": "3BR", "D": "3BR"}.get(prefix, "Unknown")


def get_wst_ner_floor(unit_type: str, market_rent: float):
    """Compute NER floor for Western Station based on floor plan and market rent.
    Per Daniel Twito: A1/A1U/A2U = $150 off market, other 1BR = $200 off,
    2BR = $300 off, 3BR = $200 off.
    """
    if not market_rent or market_rent <= 0:
        return None
    clean = unit_type.strip().upper()
    base  = re.split(r'[-\s]', clean)[0] if clean else ""
    if base in ("A1", "A1U", "A2U"):
        discount = 150
    elif clean.startswith("A"):
        discount = 200   # other 1BR
    elif clean.startswith("B"):
        discount = 300   # 2BR
    elif clean.startswith(("C", "D")):
        discount = 200   # 3BR
    else:
        return None
    return max(market_rent - discount, 0)


def make_flag(property_name: str, unit: str, resident: str,
              rule: str, detail: str, amount_impact: float,
              source_file: str) -> dict:
    """Build a standardised exception record."""
    return {
        "Property":      property_name,
        "Unit":          unit,
        "Resident":      resident,
        "Rule":          rule,
        "Risk_Level":    RISK_MAP.get(rule, RISK_MEDIUM),
        "Detail":        detail,
        "Amount_Impact": round(float(amount_impact), 2),
        "Source_File":   source_file,
    }


# ===========================================================================
# SECTION 2 -- SPECIALISED LOADERS (one per ResMan report type)
# ===========================================================================

def _csv_files(folder: str) -> list:
    if not os.path.exists(folder):
        print(f"  [WARN] Folder missing: {folder}")
        return []
    files = [f for f in os.listdir(folder) if f.lower().endswith(".csv")]
    if not files:
        print(f"  [INFO] No CSVs in: {folder}")
    return files


def _read_csv(fpath_or_buffer, **kwargs) -> pd.DataFrame:
    """Accept a file path or file-like object (BytesIO / UploadedFile). Try encodings in order."""
    if hasattr(fpath_or_buffer, "read"):
        if hasattr(fpath_or_buffer, "seek"):
            fpath_or_buffer.seek(0)
        raw = fpath_or_buffer.read()
        for enc in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                return pd.read_csv(io.BytesIO(raw), encoding=enc, **kwargs)
            except UnicodeDecodeError:
                continue
        raise ValueError("Could not decode uploaded file with utf-8-sig, cp1252, or latin-1.")
    for enc in ("utf-8-sig", "cp1252", "latin-1"):
        try:
            return pd.read_csv(fpath_or_buffer, encoding=enc, **kwargs)
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not decode {fpath_or_buffer} with utf-8-sig, cp1252, or latin-1.")


# -- 2-A  Transaction List (Credits) ----------------------------------------
def load_transaction_list(folder: str, uploaded_files=None) -> pd.DataFrame:
    """
    ResMan Transaction List format:
      Rows 1-6  : property header block (skip)
      Row 7     : column headers
      Rows 8+   : data rows interspersed with category headers
                  e.g. 'Credit - Concession - Rent'

    Columns: Date, Reference, Unit, Name, Description, Notes, Amount,
             Gross Payments, Reverse Date, In Period Reversal,
             Out Of Period Reversal, Period Charges, Prior Charges,
             Post Charges, Related

    Important:
      - File contains BOTH original credits (positive) AND their reversals
        (negative Amount, Description prefixed 'Reversed ...').
      - Net concession per unit = SUM(all amounts).
      - We keep every row that has a numeric Unit value.
    """
    all_data = []
    _sources = ([(f.name, f) for f in uploaded_files] if uploaded_files is not None
                else [(fn, os.path.join(folder, fn)) for fn in _csv_files(folder)])
    for fname, src in _sources:
        try:
            df = _read_csv(src, skiprows=6, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            df["_prop"] = derive_property(fname)
            df["Source_File"] = fname

            # Forward-fill section headers (e.g. "Credit - Concession - Rent",
            # "Charge - Rent") which ResMan writes into the first (Date) column.
            # Only keep rows that belong to a concession-relevant "Credit - " section.
            # Excluded sections:
            #   "Credit - Renters Insurance Premium Credit" â€” these are resident
            #   insurance PAYMENT rows (Description='Payment'), not concessions.
            #
            # IMPORTANT: ALL section header types must be captured in the ffill so
            # that non-Credit sections (Payment, Deposit, etc.) properly reset the
            # section label.  Without this, Payment/Deposit rows that appear after
            # a Credit section inherit the last Credit section label via ffill and
            # are wrongly treated as concession credits.
            # Covers: Payment - Payment, Payment - Payment On Behalf Of Resident,
            #         Deposit Applied to Balance - ..., Deposit Refund - ...,
            #         Deposit - Security Deposit (seen across all 7 properties).
            NON_CONCESSION_SECTIONS = {
                "Credit - Renters Insurance Premium Credit",
            }
            date_col = df.iloc[:, 0].astype(str).str.strip()
            df["_section"] = date_col.where(
                date_col.str.match(r"^(Credit|Charge|Payment|Deposit)")
            ).ffill()
            df = df[
                df["_section"].str.startswith("Credit - ", na=False) &
                ~df["_section"].isin(NON_CONCESSION_SECTIONS) &
                df["Unit"].astype(str).str.strip().str.match(r"^\d+$")
            ].copy()

            df["Property"]   = df["_prop"]
            df["Unit"]       = df["Unit"].apply(clean_unit)
            df["Amount"]     = df["Amount"].apply(clean_currency)
            df["Is_Reversal"] = df["Amount"] < 0
            df["Description"] = df.get("Description", pd.Series(dtype=str)).fillna("").str.strip()
            # Strip ResMan status markers (* = NTV, ** = MTM) from resident names.
            # Use full replace (not lstrip) because a cell can contain comma-joined names
            # where an interior name carries the marker e.g. "Jane Doe, *John Smith"
            df["Name"]        = df.get("Name", pd.Series(dtype=str)).fillna("Unknown") \
                                   .str.replace(r"\*+", "", regex=True).str.strip()
            df["Date"]        = df.get("Date", pd.Series(dtype=str)).apply(parse_date)

            rdate = df.columns[df.columns.str.strip() == "Reverse Date"]
            df["Reverse Date"] = df[rdate[0]] if len(rdate) else ""

            # Drop internal working columns before returning
            # Note: _section is kept (not dropped) so downstream engines can
            # identify credit type (Concession vs Referral vs Employee Unit).
            df = df.drop(columns=["_prop"], errors="ignore")

            all_data.append(df)
            print(f"  [OK] Transactions: {fname}  ({len(df)} rows)")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}")

    return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()


# -- 2-B  New & Renewed Leases -----------------------------------------------
def load_leases(folder: str, uploaded_files=None) -> pd.DataFrame:
    """
    ResMan New & Renewed Leases format:
      Rows 1-5  : header block (skip)
      Row 6     : column headers
      Rows 7+   : data rows with 'New Leases' / 'Renewed Leases' section headers

    Columns: Unit, Unit Type, Xfer, Residents, (blank x2), Leasing Agent,
             Application/Renewal Date, Lease Signed Date, (blank),
             Lease Start Date, Lease End Date, Prior Rent, Market Rent,
             Rent, Rec. Conc., One Time Conc., $ Change, % Change

    Rec. Conc.    = approved monthly recurring concession
    One Time Conc.= approved one-time concession
    """
    all_data = []
    _sources = ([(f.name, f) for f in uploaded_files] if uploaded_files is not None
                else [(fn, os.path.join(folder, fn)) for fn in _csv_files(folder)])
    for fname, src in _sources:
        try:
            df = _read_csv(src, skiprows=5, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            df["Property"]    = derive_property(fname)
            df["Source_File"] = fname

            # Keep only valid data rows: Unit must be a number
            df = df[df["Unit"].astype(str).str.strip().str.match(r"^\d+$")].copy()

            df["Unit"]          = df["Unit"].apply(clean_unit)
            df["Lease Start"]   = df.get("Lease Start Date", pd.Series(dtype=str)).apply(parse_date)
            df["Lease End"]     = df.get("Lease End Date",   pd.Series(dtype=str)).apply(parse_date)
            df["Rec_Conc"]      = df.get("Rec. Conc.",       pd.Series(dtype=str)).apply(clean_currency)
            df["One_Time_Conc"] = df.get("One Time Conc.",   pd.Series(dtype=str)).apply(clean_currency)
            df["Rent"]          = df.get("Rent",             pd.Series(dtype=str)).apply(clean_currency)
            df["Market_Rent"]   = df.get("Market Rent",      pd.Series(dtype=str)).apply(clean_currency)
            df["Residents"]     = df.get("Residents",        pd.Series(dtype=str)).fillna("Unknown")
            df["Leasing_Agent"] = df.get("Leasing Agent",    pd.Series(dtype=str)).fillna("Unknown")

            all_data.append(df)
            print(f"  [OK] Leases: {fname}  ({len(df)} rows)")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}")

    return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()


# -- 2-C  Edited Transactions by User ----------------------------------------
def load_edits(folder: str, uploaded_files=None) -> pd.DataFrame:
    """
    ResMan Edited Transactions by User format:
      Rows 1-5  : header block (skip)
      Row 6     : column headers
      Rows 7+   : MIXED rows:
                    Manager name row  : Date col = manager login, rest blank
                    Data row          : Date is M/D/YYYY

    Columns: Date, Reference, Unit, Name, Category, (blank), Description,
             Amount, Reversal Date, Reversal Notes, Edited Date, Edited Amount

    Two event types captured:
      Reversal     : Reversal Date populated         -> revenue_impact = -Amount
      Amount Change: Edited Amount populated != Amount -> revenue_impact = edited - orig
    """
    all_data = []
    _sources = ([(f.name, f) for f in uploaded_files] if uploaded_files is not None
                else [(fn, os.path.join(folder, fn)) for fn in _csv_files(folder)])
    for fname, src in _sources:
        try:
            df = _read_csv(src, skiprows=5, dtype=str, header=0)
            df.columns = [str(c).strip() for c in df.columns]
            prop = derive_property(fname)

            current_manager = "Unknown"
            records = []

            for _, row in df.iterrows():
                val = str(row.iloc[0]).strip()
                if val in ("nan", ""):
                    continue

                if is_date_string(val):
                    unit      = clean_unit(row.get("Unit", ""))
                    orig_amt  = clean_currency(row.get("Amount", 0))
                    rev_date  = str(row.get("Reversal Date", "")).strip()
                    edited_raw = str(row.get("Edited Amount", "")).strip()
                    edited_amt = clean_currency(edited_raw) if edited_raw not in ("", "nan") else None

                    is_reversal   = rev_date not in ("", "nan")
                    is_amt_change = (edited_amt is not None and abs(edited_amt - orig_amt) > 0.01)

                    if not is_reversal and not is_amt_change:
                        continue
                    # Skip reversals of $0 transactions â€” no real revenue impact
                    if is_reversal and orig_amt == 0.0:
                        continue

                    revenue_impact = -orig_amt if is_reversal else (edited_amt - orig_amt)
                    event_type     = "Reversal" if is_reversal else "Amount Change"

                    records.append({
                        "Property":        prop,
                        "Manager_Login":   current_manager,
                        "Unit":            unit,
                        "Resident":        re.sub(r"\*+", "", str(row.get("Name", "Unknown"))).strip(),
                        "Category":        str(row.get("Category", "Unknown")),
                        "Description":     str(row.get("Description", "")),
                        "Original_Amount": orig_amt,
                        "Edited_Amount":   edited_amt if edited_amt is not None else orig_amt,
                        "Event_Type":      event_type,
                        "Revenue_Impact":  revenue_impact,
                        "Date":            parse_date(val),
                        "Source_File":     fname,
                    })
                else:
                    if not val.startswith("Date"):
                        current_manager = val

            if records:
                all_data.append(pd.DataFrame(records))
                print(f"  [OK] Edits: {fname}  ({len(records)} events)")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}")

    return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()


# -- 2-D  Transaction Projection (Recurring) ---------------------------------
def load_transaction_projection(folder: str, audit_month: str = AUDIT_MONTH, uploaded_files=None) -> pd.DataFrame:
    """
    ResMan Transaction Projection format (3-section file):
      Section 1: 'Recurring Transactions by Unit Type'   (summary)
      Section 2: 'Recurring Transactions by Transaction Category'  (summary)
      Section 3: 'Recurring Transactions by Unit'  <-- the useful one

    After finding the Section 3 marker we read:
      Next row  = column headers: Unit, Unit type, Category, Feb 2026, Mar 2026 ...
      Data rows = '101 - Inez Lee', '1X1 B', 'Billing Fee', 5.00, 5.00, ...

    Unit format '101 - Inez Lee' -> unit='101', resident='Inez Lee'
    We extract the column matching audit_month (e.g. 'Feb 2026').
    """
    all_data = []
    _sources = ([(f.name, f) for f in uploaded_files] if uploaded_files is not None
                else [(fn, os.path.join(folder, fn)) for fn in _csv_files(folder)])
    for fname, src in _sources:
        try:
            raw  = _read_csv(src, header=None, dtype=str)
            prop = derive_property(fname)

            # Find 'Recurring Transactions by Unit' section marker
            marker_rows = raw[
                raw.iloc[:, 0].astype(str).str.strip() == "Recurring Transactions by Unit"
            ].index

            if marker_rows.empty:
                print(f"  [WARN] 'Recurring Transactions by Unit' not found in {fname}")
                continue

            section_start = marker_rows[0] + 1       # row with column headers
            header_row    = raw.iloc[section_start]

            # Find the audit month column
            month_col_idx = None
            for i, h in enumerate(header_row):
                if audit_month.lower() in str(h).lower():
                    month_col_idx = i
                    break
            if month_col_idx is None:
                month_col_idx = 3  # fall back to first month column

            # Extract data rows after header
            data = raw.iloc[section_start + 1:].copy()
            data.columns = range(len(data.columns))
            # Keep only rows where col 0 starts with a digit (unit rows)
            data = data[data[0].astype(str).str.strip().str.match(r"^\d")]

            records = []
            for _, row in data.iterrows():
                raw_unit  = str(row[0]).strip()
                unit_num  = clean_unit(raw_unit)
                # Strip ResMan status markers (* = NTV, ** = MTM) from resident name
                resident  = re.sub(r"\*+", "", raw_unit.split(" - ", 1)[1]).strip() if " - " in raw_unit else "Unknown"
                unit_type = str(row[1]).strip() if pd.notna(row[1]) else ""
                category  = str(row[2]).strip() if pd.notna(row[2]) else ""
                amount    = clean_currency(row[month_col_idx])

                records.append({
                    "Property":    prop,
                    "Unit":        unit_num,
                    "Resident":    resident,
                    "Unit_Type":   unit_type,
                    "Category":    category,
                    "Amount":      amount,
                    "Source_File": fname,
                })

            if records:
                all_data.append(pd.DataFrame(records))
                print(f"  [OK] Projection: {fname}  ({len(records)} rows)")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}")

    return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()


def load_transaction_projection_all_months(folder: str, uploaded_files=None) -> pd.DataFrame:
    """
    Like load_transaction_projection but extracts ALL month columns.
    Returns long-format DataFrame â€” one row per (unit, category, month):
      Property, Unit, Resident, Unit_Type, Category, Month_Label, Amount, Source_File

    Used by the NER engine (Stage 1B) to detect multi-month recurring anomalies:
    full-rent offsets, double-discounts, and net effective rent below property floors.
    Concession amounts will be negative (as stored by ResMan); rent amounts positive.
    """
    all_data  = []
    month_re  = re.compile(r"^[A-Z][a-z]{2} \d{4}$")
    _sources = ([(f.name, f) for f in uploaded_files] if uploaded_files is not None
                else [(fn, os.path.join(folder, fn)) for fn in _csv_files(folder)])
    for fname, src in _sources:
        try:
            raw  = _read_csv(src, header=None, dtype=str)
            prop = derive_property(fname)

            marker_rows = raw[
                raw.iloc[:, 0].astype(str).str.strip() == "Recurring Transactions by Unit"
            ].index
            if marker_rows.empty:
                continue

            section_start = marker_rows[0] + 1
            header_row    = raw.iloc[section_start]

            # Collect all month columns (anything matching "Mon YYYY")
            month_cols = {}
            for i in range(3, len(header_row)):
                h = str(header_row.iloc[i]).strip()
                if month_re.match(h):
                    month_cols[i] = h
            if not month_cols:
                continue

            data = raw.iloc[section_start + 1:].copy()
            data.columns = range(len(data.columns))
            data = data[data[0].astype(str).str.strip().str.match(r"^\d")]

            records = []
            for _, row in data.iterrows():
                raw_unit  = str(row.iloc[0]).strip()
                unit_num  = clean_unit(raw_unit)
                resident  = (re.sub(r"\*+", "", raw_unit.split(" - ", 1)[1]).strip()
                             if " - " in raw_unit else "Unknown")
                unit_type = str(row.iloc[1]).strip() if pd.notna(row.iloc[1]) else ""
                category  = str(row.iloc[2]).strip() if pd.notna(row.iloc[2]) else ""
                for col_idx, month_label in month_cols.items():
                    amount = (clean_currency(row.iloc[col_idx])
                              if col_idx < len(row) else 0.0)
                    records.append({
                        "Property":    prop,
                        "Unit":        unit_num,
                        "Resident":    resident,
                        "Unit_Type":   unit_type,
                        "Category":    category,
                        "Month_Label": month_label,
                        "Amount":      amount,
                        "Source_File": fname,
                    })
            if records:
                all_data.append(pd.DataFrame(records))
        except Exception as e:
            print(f"  [ERROR] multi-month projection {fname}: {e}")

    return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()


# -- 2-E  Rent Roll ----------------------------------------------------------
def load_rent_roll(folder: str, uploaded_files=None) -> pd.DataFrame:
    """
    ResMan Rent Roll format (multi-row per unit):
      Rows 1-6  : header block (skip)
      Row 7     : column headers (sparse, use positional indexing)
      Rows 8+   : alternating UNIT HEADER rows + CHARGE rows + TOTAL row

    Unit header row : col[0] = unit number (e.g. '101')
    Charge sub-rows : col[0] = blank, col[18] = description, col[22] = amount
    Total row       : col[18] = 'Total'  -> skip

    Column positions (0-based after skiprows=6):
      [0]  Unit number         [2]  Unit Type
      [5]  Residents           [10] Status
      [12] Market Rent         [18] Charge Description
      [22] Charge Amount       [25] Move In
      [26] Lease Start         [27] Lease End
      [36] Balance
    """
    all_data = []
    _sources = ([(f.name, f) for f in uploaded_files] if uploaded_files is not None
                else [(fn, os.path.join(folder, fn)) for fn in _csv_files(folder)])
    for fname, src in _sources:
        try:
            raw  = _read_csv(src, skiprows=6, header=0, dtype=str)
            prop = derive_property(fname)

            # Pad to at least 37 columns
            while len(raw.columns) < 37:
                raw[f"_pad_{len(raw.columns)}"] = np.nan

            current_unit        = None
            current_resident    = None
            current_type        = None
            current_status      = None
            current_mkt_rent    = 0.0
            current_move_in     = None
            current_lease_start = None
            current_lease_end   = None
            current_balance     = 0.0

            records = []

            for _, row in raw.iterrows():
                c0  = str(row.iloc[0]).strip()   # unit number or blank
                c2  = str(row.iloc[2]).strip()   # unit type
                c5  = str(row.iloc[5]).strip()   # residents
                c10 = str(row.iloc[10]).strip()  # status
                c12 = str(row.iloc[12]).strip()  # market rent
                c18 = str(row.iloc[18]).strip()  # description
                c21 = str(row.iloc[21]).strip()  # amount  (col 21, not 22)
                c25 = str(row.iloc[25]).strip()  # move in
                c26 = str(row.iloc[26]).strip()  # lease start
                c27 = str(row.iloc[27]).strip()  # lease end
                c35 = str(row.iloc[35]).strip() if len(row) > 35 else ""  # balance

                if c18.lower() == "total":
                    continue

                if re.match(r"^\d+$", c0):
                    # Unit header row â€” strip ResMan status markers (* = NTV, ** = MTM)
                    current_unit        = clean_unit(c0)
                    current_resident    = re.sub(r"\*+", "", c5).strip() if c5 not in ("", "nan") else "Unknown"
                    current_type        = c2
                    current_status      = c10
                    current_mkt_rent    = clean_currency(c12)
                    current_move_in     = parse_date(c25)
                    current_lease_start = parse_date(c26)
                    current_lease_end   = parse_date(c27)
                    current_balance     = clean_currency(c35)

                    if c18 not in ("", "nan"):
                        records.append({
                            "Property":    prop,
                            "Unit":        current_unit,
                            "Residents":   current_resident,
                            "Unit_Type":   current_type,
                            "Status":      current_status,
                            "Market_Rent": current_mkt_rent,
                            "Description": c18,
                            "Amount":      clean_currency(c21),
                            "Move_In":     current_move_in,
                            "Lease_Start": current_lease_start,
                            "Lease_End":   current_lease_end,
                            "Balance":     current_balance,
                            "Source_File": fname,
                        })

                elif current_unit and c18 not in ("", "nan"):
                    # Charge sub-row for current unit
                    records.append({
                        "Property":    prop,
                        "Unit":        current_unit,
                        "Residents":   current_resident,
                        "Unit_Type":   current_type,
                        "Status":      current_status,
                        "Market_Rent": current_mkt_rent,
                        "Description": c18,
                        "Amount":      clean_currency(c21),
                        "Move_In":     current_move_in,
                        "Lease_Start": current_lease_start,
                        "Lease_End":   current_lease_end,
                        "Balance":     current_balance,
                        "Source_File": fname,
                    })

            if records:
                all_data.append(pd.DataFrame(records))
                print(f"  [OK] Rent Roll: {fname}  ({len(records)} charge rows)")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}")

    return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()


# -- 2-F  Resident Activity --------------------------------------------------
def load_resident_activity(folder: str, uploaded_files=None) -> pd.DataFrame:
    """
    ResMan Resident Activity format (very wide, ~73 columns):
      Rows 1-6  : header block (skip)
      Row 7     : sparse column headers
      Row 8     : 'Adjusted Lease End Date' overflow header -- skip
      Rows 9+   : data (one row per lease record)

    Key column positions (0-based after skiprows=6):
      [0]  Unit           [2]  Residents
      [18] Unit Type      [23] Actual Rent
      [29] Move In        [32] Initial Lease End
      [37] Lease Start    [43] Lease End
      (last non-blank after col 43 = Manager)
    """
    all_data = []
    _sources = ([(f.name, f) for f in uploaded_files] if uploaded_files is not None
                else [(fn, os.path.join(folder, fn)) for fn in _csv_files(folder)])
    for fname, src in _sources:
        try:
            raw  = _read_csv(src, skiprows=6, header=0, dtype=str)
            prop = derive_property(fname)

            # Drop the 'Adjusted Lease End Date' overflow header row
            raw = raw[~raw.iloc[:, 0].astype(str).str.contains("Adjusted", na=False)]

            # Keep only unit-number rows
            raw = raw[raw.iloc[:, 0].astype(str).str.strip().str.match(r"^\d+$")]

            while len(raw.columns) < 50:
                raw[f"_pad_{len(raw.columns)}"] = np.nan

            records = []
            for _, row in raw.iterrows():
                unit        = clean_unit(str(row.iloc[0]))
                residents   = str(row.iloc[2]).strip()
                unit_type   = str(row.iloc[18]).strip() if len(row) > 18 else ""
                actual_rent = clean_currency(row.iloc[23]) if len(row) > 23 else 0.0
                move_in     = parse_date(row.iloc[29]) if len(row) > 29 else None
                lease_start = parse_date(row.iloc[37]) if len(row) > 37 else None
                lease_end   = parse_date(row.iloc[43]) if len(row) > 43 else None

                manager = "Unknown"
                for i in range(len(row) - 1, 43, -1):
                    v = str(row.iloc[i]).strip()
                    if v not in ("", "nan"):
                        manager = v
                        break

                records.append({
                    "Property":    prop,
                    "Unit":        unit,
                    "Residents":   residents,
                    "Unit_Type":   unit_type,
                    "Actual_Rent": actual_rent,
                    "Move_In":     move_in,
                    "Lease_Start": lease_start,
                    "Lease_End":   lease_end,
                    "Manager":     manager,
                    "Source_File": fname,
                })

            if records:
                all_data.append(pd.DataFrame(records))
                print(f"  [OK] Activity: {fname}  ({len(records)} rows)")
        except Exception as e:
            print(f"  [ERROR] {fname}: {e}")

    return pd.concat(all_data, ignore_index=True) if all_data else pd.DataFrame()


# -- 2-G  Market Rent Schedule Detail ----------------------------------------
def load_market_rent_schedule(folder: str, uploaded_files=None) -> dict:
    """
    ResMan Market Rent Schedule Detail format:
      Rows 1-5  : header block (property name, company, report title, date, printed)
      Row 6     : column headers:
                    Unit, Amenities, [blank], Apts. Sq. Ft., Base Mkt. Rent,
                    Base Mkt./SF, Premium, Total Rent, Total Rent/SF, Cumulative Rent
      Rows 7+   : data rows interspersed with unit-type section labels and totals.
                  - Unit-type label rows : first column is a floor-plan code like
                    "A1", "A1-P", "B2", "1X1", "2X2" etc. (not a unit number)
                  - Unit data rows       : first column is a numeric unit number
                    (3- or 4-digit, e.g. "105" or "0105")
                  - Total rows           : first column = "Total"

    Returns:
      dict mapping (property_key, unit_number_str) -> float total_market_rent
      (the "Total Rent" column, which includes all premiums)

    Used by the Revenue Integrity Engine to compute dynamic NER floors for properties where
    the floor is defined as (market_rent - PROPERTY_NER_DISCOUNT).
    """
    lookup = {}
    _sources = ([(f.name, f) for f in uploaded_files] if uploaded_files is not None
                else [(fn, os.path.join(folder, fn)) for fn in _csv_files(folder)])
    for fname, src in _sources:
        try:
            prop = derive_property(fname)
            df   = _read_csv(src, skiprows=5, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]

            # Keep only rows where the first column is a numeric unit number
            unit_col = df.iloc[:, 0].astype(str).str.strip()
            unit_mask = unit_col.str.match(r"^\d{3,4}$")
            df = df[unit_mask].copy()

            # "Total Rent" is the 8th column (index 7): base + premium
            total_rent_col = df.columns[7] if len(df.columns) > 7 else None
            if total_rent_col is None:
                print(f"  [WARN] Market Rent Schedule: {fname} has unexpected columns.")
                continue

            for _, row in df.iterrows():
                unit        = clean_unit(str(row.iloc[0]).strip())
                total_rent  = clean_currency(row[total_rent_col])
                if total_rent and total_rent > 0:
                    lookup[(prop, unit)] = total_rent

            print(f"  [OK] Market Rent Schedule: {fname}  ({len(df)} units)")
        except Exception as e:
            print(f"  [ERROR] Market Rent Schedule {fname}: {e}")

    return lookup


# ===========================================================================
# SECTION 3 -- CONCESSION AUDIT ENGINE (9 Concession Rules)
# ===========================================================================

def run_concession_audit_engine(df_trans: pd.DataFrame, df_leases: pd.DataFrame,
                     df_projection: pd.DataFrame = None,
                     df_rent_roll: pd.DataFrame = None) -> pd.DataFrame:
    """
    Concession audit rules.

    Rules implemented
    -----------------
    R1  Post-Term Credit           : credit posted after lease end date (CRITICAL)
    R2  Missing Lease              : credit posted but unit has no active lease on the
                                         Rent Roll (Lease_End is in the past or absent) (HIGH)
    R3  Large Credit (>=$700)      : any single credit >= $700 (CRITICAL)
    R4  Non-Standard Description   : freeform description, no approved keyword (MEDIUM)
    R5  Missing Addendum / No RR Setup : credit posted but NO concession row on
                                          Rent Roll for that unit (CRITICAL)
    R6  Amount Mismatch            : Rent Roll concession amount != Transaction List
                                     credit amount (>10% or >$10 delta) (HIGH)
    R7  Not Properly Posted        : Rent Roll has concession setup but no credit
                                     was posted in Transaction List (HIGH)

    NOTE: Rules 5/6/7 use the Rent Roll concession rows (negative amount rows
    with concession keywords in Description) as the "approved concession" source.
    This is the LiveNjoy data reality â€” approved concessions are stored as discounted
    rent line items on the Rent Roll, not in a separate Rec_Conc field.
    """
    print("\n[CONCESSION AUDIT ENGINE] Running concession audit rules...")
    flags = []

    # Only skip if BOTH transaction list AND rent roll have no data.
    # R7 (Not Properly Posted) works from the Rent Roll side â€” it flags units
    # where the Rent Roll shows a concession setup but no credit was posted in
    # the Transaction List. When the TX list is empty (e.g. no Credit sections
    # were exported for the month), R7 should still fire for all RR concession
    # units. R1â€“R4 (which need TX rows) are simply skipped when df_trans is empty.
    no_tx   = df_trans is None or df_trans.empty
    no_rr   = df_rent_roll is None or df_rent_roll.empty
    if no_tx and no_rr:
        print("  [SKIP] No transaction data and no rent roll data.")
        return pd.DataFrame()
    if no_tx:
        print("  [WARN] No Credit rows in Transaction List â€” R1/R2/R3/R4 skipped. "
              "R7 (Not Properly Posted) will still run from Rent Roll.")

    # Keywords that identify a concession/discount row on the Rent Roll
    CONC_RR_KW = [
        "concession", "$999", "special", "reduce", "employee", "discount",
        "free", "$200", "$100", "concession rent", "allowance", "courtesy",
        "mi special", "move in", "move-in", "rent concession",
    ]

    # Standard description keywords present in approved concession descriptions
    CONC_TX_KW = [
        "concession", "allowance", "employee unit", "courtesy officer",
        "resident referral", "referral", "move in special", "move-in special",
        "reduce", "special", "discount",
    ]

    # ------------------------------------------------------------------
    # Build lookups
    # ------------------------------------------------------------------

    # Lease lookup: (prop, unit) -> most recent lease row from New & Renewed Leases
    # Used only for R1 (Post-Term Credit) date comparison.
    lease_lookup = {}
    if df_leases is not None and not df_leases.empty:
        for _, row in df_leases.sort_values("Lease Start", ascending=False).iterrows():
            key = (row["Property"], row["Unit"])
            if key not in lease_lookup:
                lease_lookup[key] = row

    # Rent Roll lease-end lookup: (prop, unit) -> lease_end date
    # Used for R2 (Missing Lease) â€” checks whether a unit currently has an active
    # lease. This is far more reliable than checking only new/renewed leases for
    # the current month, which would miss all ongoing leases signed in prior months.
    rr_lease_end_lookup   = {}
    rr_lease_start_lookup = {}
    if df_rent_roll is not None and not df_rent_roll.empty:
        for (prop, unit), grp in df_rent_roll.groupby(["Property", "Unit"]):
            lease_end_val   = grp["Lease_End"].dropna().max()
            lease_start_val = grp["Lease_Start"].dropna().min()
            rr_lease_end_lookup[(prop, unit)]   = lease_end_val
            rr_lease_start_lookup[(prop, unit)] = lease_start_val

    # Rent Roll concession lookup: (prop, unit) -> approved monthly concession $ (positive)
    # Only negative-amount rows matching concession keywords are counted
    rr_conc_lookup = {}   # (prop, unit) -> abs(sum of concession amounts)
    rr_src_lookup  = {}   # (prop, unit) -> source file
    rr_conc_desc_lookup = {}  # (prop, unit) -> first concession description on Rent Roll
    if df_rent_roll is not None and not df_rent_roll.empty:
        conc_mask = df_rent_roll["Description"].str.lower().apply(
            lambda x: any(k in x for k in CONC_RR_KW)
        ) & (df_rent_roll["Amount"] < -0.01)   # only actual monetary reductions
        rr_conc_rows = df_rent_roll[conc_mask]
        for (prop, unit), grp in rr_conc_rows.groupby(["Property", "Unit"]):
            rr_conc_lookup[(prop, unit)] = abs(grp["Amount"].sum())
            rr_src_lookup[(prop, unit)]  = grp["Source_File"].iloc[0]
            rr_conc_desc_lookup[(prop, unit)] = str(grp["Description"].iloc[0])

    # Transaction List credit lookup: (prop, unit) -> NET credit after reversals
    # Per Daniel Twito (June 4 meeting): the bot must use the NET of all concession
    # activity â€” positive credits minus reversed credits â€” not just the sum of
    # positives. A $346 credit that was then reversed and replaced with a $500
    # credit should show as $500 net, not $846 total positives.
    _R5_SKIP_SECTIONS = {"Credit - Resident Referral", "Credit - Employee Unit Rent Allowance"}
    tx_credit_lookup   = {}   # (prop, unit) -> NET credit (after reversals)
    tx_credits_detail  = {}   # (prop, unit) -> list of (desc, amount) for each credit row
    tx_src_lookup      = {}
    tx_res_lookup      = {}
    if not df_trans.empty:
        _conc_desc_mask = df_trans["Description"].str.lower().apply(
            lambda x: any(k in x for k in CONC_TX_KW)
        )
        conc_rows = df_trans[
            (df_trans["Amount"].abs() > 0) &
            (~df_trans["_section"].isin(_R5_SKIP_SECTIONS)) &
            _conc_desc_mask
        ]
        for (prop, unit), grp in conc_rows.groupby(["Property", "Unit"]):
            # Net = sum of all credit amounts (positive = credit, negative = reversal)
            # Reversals are stored as negative amounts in the transaction list.
            net = grp["Amount"].sum()
            # Always add to lookup if concession activity exists (even if net=0 due to
            # full reversal) so R7 (Not Properly Posted) doesn't fire â€” R6 (Concession
            # Amount Mismatch) will correctly catch the $0 posted vs RR approved amount.
            if net >= 0:
                tx_credit_lookup[(prop, unit)] = net
                tx_src_lookup[(prop, unit)]    = grp["Source_File"].iloc[0]
                tx_res_lookup[(prop, unit)]    = grp["Name"].iloc[0]
                tx_credits_detail[(prop, unit)] = [
                    (str(r["Description"]), float(r["Amount"]))
                    for _, r in grp.iterrows()
                ]

    # Rent Roll market rent lookup (for R5-Missing-Addendum context)
    rr_market_lookup = {}
    if df_rent_roll is not None and not df_rent_roll.empty:
        for (prop, unit), grp in df_rent_roll.groupby(["Property", "Unit"]):
            rr_market_lookup[(prop, unit)] = grp["Market_Rent"].iloc[0]

    # ------------------------------------------------------------------
    # R1, R2, R3, R4 â€” per transaction-list unit loop
    # (skipped when no Credit rows were exported from ResMan)
    # ------------------------------------------------------------------
    for (prop, unit), grp in (df_trans.groupby(["Property", "Unit"]) if not no_tx else []):
        resident  = grp["Name"].iloc[0]
        src       = grp["Source_File"].iloc[0]
        lease_row = lease_lookup.get((prop, unit))
        lease_end = lease_row.get("Lease End") if lease_row is not None else None

        active_credits = grp[~grp["Is_Reversal"] & (grp["Amount"] > 0)]
        net_actual     = grp["Amount"].sum()

        # R1 â€” Post-Term Credit
        if lease_end is not None and pd.notna(lease_end):
            post_term = active_credits[
                active_credits["Date"].notna() &
                (active_credits["Date"] > pd.Timestamp(lease_end))
            ]
            for _, row in post_term.iterrows():
                flags.append(make_flag(prop, unit, resident, "Post-Term Credit",
                    f"${row['Amount']:.2f} credit posted on {row['Date'].date()} "
                    f"after lease end {pd.Timestamp(lease_end).date()}. "
                    f"Description: '{row['Description']}'.",
                    row["Amount"], src))

        # R2 â€” Missing Lease
        # A unit is considered to have an active lease if its Lease_End on the
        # Rent Roll is today or in the future (or is missing/blank, which means
        # MTM or no end date set). Flag only if Lease_End is clearly in the past.
        rr_lease_end = rr_lease_end_lookup.get((prop, unit))
        rr_lease_start = rr_lease_start_lookup.get((prop, unit))
        today = pd.Timestamp.today().normalize()
        has_active_lease = (
            rr_lease_end is None
            or pd.isna(rr_lease_end)
            or pd.Timestamp(rr_lease_end) >= today
        )
        if net_actual > 0 and not has_active_lease:
            lease_end_str = pd.Timestamp(rr_lease_end).date() if rr_lease_end and pd.notna(rr_lease_end) else "unknown"
            lease_start_str = pd.Timestamp(rr_lease_start).date() if rr_lease_start and pd.notna(rr_lease_start) else "unknown"
            flags.append(make_flag(prop, unit, resident, "Missing Lease",
                f"${net_actual:.2f} net credit posted but Unit {unit} at {prop} "
                f"has no active lease on the Rent Roll "
                f"(Lease: {lease_start_str} â€“ {lease_end_str}). Cannot verify authorization.",
                net_actual, src))

        # R3 â€” Large Credit (>= $700 single transaction)
        for _, row in active_credits[active_credits["Amount"] >= CONCESSION_CRITICAL_AMT].iterrows():
            flags.append(make_flag(prop, unit, resident, "Recurring Concession >$700",
                f"Single credit of ${row['Amount']:.2f} exceeds $700 threshold. "
                f"Description: '{row['Description']}'. Requires VP approval.",
                row["Amount"], src))

        # R4 â€” Non-Standard Description: DISABLED (March 10, 2026)
        # John confirmed descriptions are freeform identifiers with no standard rule.
        # Approved codes (CONR, CRTCO, EMPL, MCCR, RRFee) are confirmed but there is
        # no separate code column in the ResMan export to check against.

    # ------------------------------------------------------------------
    # R5, R6, R7 â€” cross-check Rent Roll concession setup vs Transaction List
    # ------------------------------------------------------------------
    all_units = set(list(rr_conc_lookup.keys()) + list(tx_credit_lookup.keys()))

    for (prop, unit) in all_units:
        in_rr = (prop, unit) in rr_conc_lookup
        in_tx = (prop, unit) in tx_credit_lookup

        approved_amt = rr_conc_lookup.get((prop, unit), 0.0)
        posted_amt   = tx_credit_lookup.get((prop, unit), 0.0)
        resident     = tx_res_lookup.get((prop, unit),
                       df_rent_roll[
                           (df_rent_roll["Property"] == prop) &
                           (df_rent_roll["Unit"] == unit)
                       ]["Residents"].iloc[0]
                       if df_rent_roll is not None and not df_rent_roll.empty
                       and len(df_rent_roll[(df_rent_roll["Property"] == prop) &
                                            (df_rent_roll["Unit"] == unit)]) > 0
                       else "Unknown")
        src = tx_src_lookup.get((prop, unit), rr_src_lookup.get((prop, unit), ""))
        market = rr_market_lookup.get((prop, unit), 0.0)

        # R5 â€” Missing Addendum: DISABLED (April 30, 2026)
        # Addenda are stored as PDFs in ResMan's Documents section, which is not
        # accessible from CSV exports. The Rent Roll concession row is not a
        # reliable proxy for addendum existence â€” John confirmed all flagged
        # addenda exist. This rule generated only false positives.
        # if in_tx and not in_rr:
        #     ...

        # R6 â€” Amount Mismatch: both exist but net amounts differ
        # Uses NET posted amount (after reversals) per Daniel Twito June 4 meeting.
        if in_tx and in_rr:
            delta    = abs(posted_amt - approved_amt)
            pct_diff = delta / approved_amt if approved_amt > 0 else 0
            if delta > 10 and pct_diff > 0.10:
                # Build a human-readable breakdown of the credit activity
                detail_rows = tx_credits_detail.get((prop, unit), [])
                if detail_rows:
                    parts   = []
                    for desc, amt in detail_rows:
                        sign = "+" if amt > 0 else ""
                        parts.append(f"{sign}${amt:.0f} ({desc[:80]})")
                    activity = "  |  ".join(parts)
                    if posted_amt == 0:
                        # All concession activity was reversed â€” net $0 posted
                        detail_text = (
                            f"Rent Roll has a -${approved_amt:.0f}/mo concession setup, "
                            f"but all posted credits were reversed: {activity}. "
                            f"Net posted: $0. "
                            f"Concession was NOT applied for {AUDIT_MONTH}. "
                            f"Confirm whether this was intentional or a missed posting."
                        )
                    else:
                        direction = "higher" if posted_amt > approved_amt else "lower"
                        detail_text = (
                            f"Rent Roll setup: -${approved_amt:.0f}/mo concession. "
                            f"Transaction activity: {activity}. "
                            f"Net posted: ${posted_amt:.0f} â€” ${delta:.0f} "
                            f"{direction} than Rent Roll ({pct_diff*100:.0f}% variance). "
                            f"Verify signed addendum authorizes the net amount posted."
                        )
                else:
                    direction = "higher" if posted_amt > approved_amt else "lower"
                    detail_text = (
                        f"Rent Roll has a -${approved_amt:.0f}/mo concession setup, "
                        f"but ${posted_amt:.0f} was posted to the Transaction List "
                        f"(${delta:.0f} {direction}, {pct_diff*100:.0f}% variance). "
                        f"Verify signed addendum authorizes the net amount posted."
                    )
                flags.append(make_flag(prop, unit, resident, "Concession Amount Mismatch",
                    detail_text, delta, src))

        # R7 â€” Not Properly Posted: concession on Rent Roll but nothing posted
        elif in_rr and not in_tx:
            rr_desc = rr_conc_desc_lookup.get((prop, unit), "")
            desc_note = f" (Rent Roll entry: '{rr_desc}')" if rr_desc else ""
            flags.append(make_flag(prop, unit, resident, "Not Properly Posted",
                f"Rent Roll shows a -${approved_amt:.0f}/mo concession{desc_note}, "
                f"but no matching credit was posted to the Transaction List for "
                f"{AUDIT_MONTH}. Confirm whether this is a missed posting, a "
                f"prorated move-in/out, or an NTV situation.",
                approved_amt, src))

    result = pd.DataFrame(flags) if flags else pd.DataFrame()
    print(f"  [CONCESSION AUDIT ENGINE] Complete -- {len(result)} flags.")
    return result


# ===========================================================================
# SECTION 4 -- REVENUE INTEGRITY ENGINE (2-Stage Revenue Integrity)
# ===========================================================================

def run_revenue_integrity_engine(
    df_projection: pd.DataFrame,
    df_rent_roll:  pd.DataFrame,
    df_trans:      pd.DataFrame,
    df_leases:     pd.DataFrame,
    df_activity:   pd.DataFrame,
    df_proj_full:  pd.DataFrame = None,
    market_rent_lookup: dict = None,
) -> pd.DataFrame:
    """
    Stage 1 -- Recurring Projection Audit (Transaction Projection file)
      3.1  Standard charge completeness (90% Rule)
      3.2  Amount consistency check
      3.3  Recurring concession red flags

    Stage 2 -- Posted Rent Roll Audit (Rent Roll file)
      4.1  Net rent integrity (negative or $0)
      4.2  Manual concession detection (posted but no recurring setup)
      4.3  Posted vs recurring mismatch
      4.4  Misc tenant credit review
    """
    print("\n[REVENUE INTEGRITY ENGINE] Running 2-stage revenue integrity audit...")
    flags = []

    # =========================================================================
    # STAGE 1 -- RECURRING PROJECTION
    # =========================================================================
    if not df_projection.empty:
        print("  [STAGE 1] Recurring Transaction Projection ...")
        proj = df_projection.copy()
        proj["Cat_Lower"] = proj["Category"].str.lower()

        # 3.1 -- 90% RULE: every charge standard at >= 90% of units must exist for ALL
        for prop, prop_grp in proj.groupby("Property"):
            total_units = prop_grp["Unit"].nunique()
            if total_units == 0:
                continue

            for category, cat_grp in prop_grp.groupby("Category"):
                if not category:
                    continue
                # Skip optional charge types â€” parking, pet fees, W/D are unit-specific
                if any(kw in category.lower() for kw in OPTIONAL_CHARGE_KEYWORDS):
                    continue
                units_with = cat_grp[cat_grp["Amount"] > 0]["Unit"].nunique()
                pct = units_with / total_units

                if pct >= STANDARD_CHARGE_THRESHOLD:
                    all_units  = set(prop_grp["Unit"].unique())
                    have_units = set(cat_grp[cat_grp["Amount"] > 0]["Unit"].unique())
                    missing    = all_units - have_units
                    std_amount = cat_grp[cat_grp["Amount"] > 0]["Amount"].mode()
                    std_amount = std_amount.iloc[0] if not std_amount.empty else 0.0

                    for mu in missing:
                        sub = prop_grp[prop_grp["Unit"] == mu]
                        resident = sub["Resident"].iloc[0] if not sub.empty else "Unknown"
                        src      = sub["Source_File"].iloc[0] if not sub.empty else ""
                        flags.append(make_flag(prop, mu, resident,
                            "Missing Standard Charge",
                            f"'{category}' is standard at {prop} "
                            f"({pct*100:.0f}% of units, ${std_amount:.2f}/mo). "
                            f"Unit {mu} has no charge set.",
                            std_amount, src))

        # 3.2 -- AMOUNT CONSISTENCY: flag within the same Property + Unit Type + Category
        # (comparing same unit type only avoids false positives across 1BR vs 2BR etc.)
        # NOTE: "Rent" is excluded from Minor variance â€” rent legitimately varies between
        # units of the same type (different lease terms, negotiated rates). Significant rent
        # outliers (>=20% and >=$5) are still caught by Major variance. Rent mismatches
        # against the Rent Roll are fully covered by Stage 2 "Posted vs Recurring Mismatch".
        RENT_KEYWORDS = {"rent"}
        for (prop, unit_type, category), grp in proj.groupby(["Property", "Unit_Type", "Category"]):
            if not category or not unit_type:
                continue
            active = grp[grp["Amount"] > 0]
            if len(active) < 3:
                continue
            mode_val = active["Amount"].mode()
            if mode_val.empty or mode_val.iloc[0] == 0:
                continue
            mode_amt = mode_val.iloc[0]

            is_rent = category.lower().strip() in RENT_KEYWORDS
            is_optional = any(kw in category.lower() for kw in OPTIONAL_CHARGE_KEYWORDS)
            for _, row in active.iterrows():
                var     = abs(row["Amount"] - mode_amt)
                pct_var = var / mode_amt
                if var >= 1.0:
                    is_major = pct_var >= 0.20 and var >= 5.0
                    # Skip rent variance entirely â€” already covered by Stage 2
                    # "Posted vs Recurring Mismatch" which compares against actual Rent Roll
                    if is_rent:
                        continue
                    # Skip optional charges â€” Fee Schedule Check owns their amount validation
                    if is_optional:
                        continue
                    rule = "Major Charge Amount Variance" if is_major \
                           else "Minor Charge Amount Variance"
                    flags.append(make_flag(prop, row["Unit"], row["Resident"], rule,
                        f"'{category}' ({unit_type}): Unit ${row['Amount']:.2f} vs "
                        f"standard ${mode_amt:.2f} for same unit type "
                        f"(Delta ${var:.2f}, {pct_var*100:.0f}%)",
                        var, row.get("Source_File", "")))

        # 3.3 -- RECURRING CONCESSION RED FLAGS
        conc_keywords = ["concession", "conr", "crtco", "empl", "mccr", "rrfee",
                         "employee unit", "resident referral", "courtesy officer"]
        conc_mask = proj["Cat_Lower"].apply(
            lambda x: any(k in x for k in conc_keywords)
        )
        conc_proj = proj[conc_mask].copy()

        if not conc_proj.empty:
            for (prop, unit), grp in conc_proj.groupby(["Property", "Unit"]):
                # Concessions are stored as negative in the projection DataFrame;
                # use abs() so threshold comparisons work correctly.
                amt      = grp["Amount"].abs().max()
                months   = (grp["Amount"].abs() > 0).sum()
                resident = grp["Resident"].iloc[0]
                src      = grp["Source_File"].iloc[0]

                if amt > CONCESSION_CRITICAL_AMT:
                    # Skip full-rent-offset concessions — Stage 1B already flags
                    # these as Full Rent Recurring Offset (CRITICAL) or Future Month
                    # Full Offset (VERIFY). Emitting a duplicate HIGH here is noise.
                    rent_for_unit = proj[
                        (proj["Property"] == prop) &
                        (proj["Unit"]     == unit) &
                        proj["Cat_Lower"].str.contains(r"\brent\b", regex=True) &
                        ~proj["Cat_Lower"].str.contains("concession") &
                        (proj["Amount"] > 0)
                    ]["Amount"].sum()
                    if rent_for_unit > 0 and abs(amt - rent_for_unit) < 15:
                        continue  # Full-rent offset handled by Stage 1B
                    flags.append(make_flag(prop, unit, resident,
                        "Recurring Concession >$700",
                        f"Recurring concession -${amt:.0f}/mo in {AUDIT_MONTH}. "
                        f"Single monthly concession exceeds $700 — requires VP approval.",
                        amt, src))

                elif abs(amt - 500) < 1.0:
                    flags.append(make_flag(prop, unit, resident,
                        "Exact $500 Concession",
                        f"Recurring concession exactly -$500/mo. "
                        f"May be $500 first full month / $500 last month split. "
                        f"Verify concession addendum is on file.",
                        amt, src))

                elif months > CONCESSION_HIGH_MONTHS and amt > CONCESSION_HIGH_AMT:
                    flags.append(make_flag(prop, unit, resident,
                        "Concession >$500 for 2+ Months",
                        f"-${amt:.0f}/mo concession for {months} month(s) in recurring setup. "
                        f"Recurring concession above ${CONCESSION_HIGH_AMT} for more than "
                        f"{CONCESSION_HIGH_MONTHS} months â€” verify lease/addendum.",
                        amt, src))

    # =========================================================================
    # STAGE 1B -- NET EFFECTIVE RENT ANALYSIS  (Daniel Twito, June 2026)
    # Uses multi-month projection to detect:
    #   â€¢ Full rent recurring offsets (concession = rent for 2+ months) â†’ CRITICAL
    #   â€¢ Single future-month full offset (one-month-free)             â†’ HIGH / VERIFY
    #   â€¢ Double-discount setup (rent reduced + concession on top)     â†’ CRITICAL
    #   â€¢ Net effective rent below property floor                      â†’ CRITICAL
    # =========================================================================
    if df_proj_full is not None and not df_proj_full.empty:
        _conc_kw = {"concession", "conr", "crtco", "empl", "mccr", "rrfee",
                    "employee unit", "resident referral", "courtesy officer"}
        pf = df_proj_full.copy()
        pf["_cat"]     = pf["Category"].str.lower().str.strip()
        pf["_is_rent"] = (
            pf["_cat"].str.contains(r"\brent\b", na=False, regex=True) &
            ~pf["_cat"].str.contains("concession", na=False)
        )
        pf["_is_conc"] = pf["_cat"].apply(lambda x: any(k in x for k in _conc_kw))

        def _month_key(m):
            try:
                return pd.to_datetime(m, format="%b %Y")
            except Exception:
                return pd.Timestamp("2099-01-01")

        all_months_sorted = sorted(pf["Month_Label"].unique(), key=_month_key)

        # Track units already flagged for NER so we don't emit both
        # Double-Discount and NER-Below-Floor for the same unit.
        # Build rent-roll status lookup for NTV detection in rent decrease audit
        rr_status_lookup = {}
        if df_rent_roll is not None and not df_rent_roll.empty:
            for (p, u), grp in df_rent_roll.groupby(["Property", "Unit"]):
                rr_status_lookup[(p, u)] = grp["Status"].iloc[0]

        _ner_flagged = set()

        for (prop, unit), ug in pf.groupby(["Property", "Unit"]):
            resident  = ug["Resident"].iloc[0]
            unit_type = ug["Unit_Type"].iloc[0]
            src       = ug["Source_File"].iloc[0]
            br_type   = get_bedroom_type(unit_type)
            ner_floor = PROPERTY_NER_FLOORS.get(prop, {}).get(br_type)

            # For properties with a market-relative floor (PROPERTY_NER_DISCOUNT),
            # compute the floor dynamically from the per-unit market rent schedule.
            # Floor = unit_market_rent - allowed_discount.
            if ner_floor is None and market_rent_lookup:
                discount = PROPERTY_NER_DISCOUNT.get(prop)
                if discount is not None:
                    sched_mkt = market_rent_lookup.get((prop, unit))
                    if sched_mkt and sched_mkt > 0:
                        ner_floor = sched_mkt - discount

            # Western Station: floor plan-specific discount from market rent
            if ner_floor is None and prop == "Western Station" and market_rent_lookup:
                sched_mkt = market_rent_lookup.get((prop, unit))
                if sched_mkt and sched_mkt > 0:
                    ner_floor = get_wst_ner_floor(unit_type, sched_mkt)

            rent_by_m = {}
            conc_by_m = {}
            for month in all_months_sorted:
                mg = ug[ug["Month_Label"] == month]
                rent_by_m[month] = mg[mg["_is_rent"]]["Amount"].sum()
                conc_by_m[month] = mg[mg["_is_conc"]]["Amount"].sum()  # stored negative

            active_months = [m for m in all_months_sorted if rent_by_m.get(m, 0) > 0]
            if not active_months:
                continue

            # â”€â”€ Full-rent offset: concession = rent â†’ NER â‰ˆ $0 â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
            full_offset_months = [
                m for m in active_months
                if abs(rent_by_m[m] + conc_by_m.get(m, 0)) < 1.0
            ]
            if len(full_offset_months) >= 2:
                first_m = full_offset_months[0]
                last_m  = full_offset_months[-1]
                rent_a  = rent_by_m[first_m]
                conc_a  = conc_by_m.get(first_m, 0)
                span    = (f"{first_m} \u2013 {last_m}"
                           if first_m != last_m else first_m)
                total_exposure = rent_a * len(full_offset_months)
                flags.append(make_flag(prop, unit, resident,
                    "Full Rent Recurring Offset",
                    f"Rent ${rent_a:.0f} / Concession -${abs(conc_a):.0f} / Net $0 \u2014 "
                    f"{span} ({len(full_offset_months)} months). "
                    f"Revenue exposure: ${total_exposure:,.0f}. "
                    f"If a one-month-free special was intended, recurring setup must be corrected to one month only.",
                    rent_a, src))
                _ner_flagged.add((prop, unit))

            elif len(full_offset_months) == 1:
                m      = full_offset_months[0]
                rent_a = rent_by_m[m]
                conc_a = conc_by_m.get(m, 0)
                # Distinguish current-month free from a genuinely upcoming free month
                if m == AUDIT_MONTH:
                    offset_context = (
                        f"{m} (current audit month). "
                        f"One-month-free concession — verify signed addendum is on file."
                    )
                else:
                    offset_context = (
                        f"{m} — one-month-free concession for upcoming month. "
                        f"Confirm signed addendum is on file before this month posts."
                    )
                flags.append(make_flag(prop, unit, resident,
                    "Future Month Full Offset",
                    f"Rent ${rent_a:.0f} / Concession -${abs(conc_a):.0f} / Net $0 \u2014 "
                    f"{offset_context}",
                    0.0, src))

            # â”€â”€ NER below floor for current audit month â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€
            if (ner_floor is not None
                    and AUDIT_MONTH in rent_by_m
                    and (prop, unit) not in _ner_flagged):
                rent_a = rent_by_m[AUDIT_MONTH]
                conc_a = conc_by_m.get(AUDIT_MONTH, 0)
                ner    = rent_a + conc_a   # conc is negative â†’ correct NER
                if rent_a > 0 and 0 < ner < ner_floor:
                    # Proration guard: skip if rent is < 40% of market rent
                    # (partial-month move-in/move-out prorations cause false positives)
                    mkt_rent = 0.0
                    if df_rent_roll is not None and not df_rent_roll.empty:
                        rr_sub = df_rent_roll[
                            (df_rent_roll["Property"] == prop) &
                            (df_rent_roll["Unit"] == unit)
                        ]
                        if not rr_sub.empty:
                            mkt_rent = rr_sub["Market_Rent"].iloc[0]
                    if mkt_rent > 0 and rent_a < mkt_rent * 0.50:
                        continue  # likely proration, not a recurring setup issue

                    is_double = (mkt_rent > 0 and rent_a < mkt_rent
                                 and abs(conc_a) > 1.0)
                    if is_double:
                        flags.append(make_flag(prop, unit, resident,
                            "Double-Discount Setup",
                            f"{unit_type} market rent ${mkt_rent:.0f}. "
                            f"Recurring setup: Rent ${rent_a:.0f} / "
                            f"Concession \u2212${abs(conc_a):.0f} / "
                            f"Net ${ner:.0f}. "
                            f"Rent is already below market AND a concession is "
                            f"applied on top \u2014 double-discount. "
                            f"{br_type} NER floor: ${ner_floor:.0f}. "
                            f"Obtain VP authorization or correct so NER \u2265 ${ner_floor:.0f}.",
                            ner_floor - ner, src))
                    else:
                        flags.append(make_flag(prop, unit, resident,
                            "Net Effective Rent Below Floor",
                            f"{unit_type} Rent ${rent_a:.0f} / Concession -${abs(conc_a):.0f} "
                            f"/ Net ${ner:.0f}. Net effective rent is below the "
                            f"{br_type} floor of ${ner_floor:.0f} for {prop}. "
                            f"Confirm authorization or correct recurring setup.",
                            ner_floor - ner, src))
                    _ner_flagged.add((prop, unit))

            # ── Rent Decrease Audit ──────────────────────────────────────────
            # Flag material recurring rent decreases from the prior month.
            # Classifies NTV units as VERIFY, all others as HIGH.
            if AUDIT_MONTH in all_months_sorted:
                audit_idx = all_months_sorted.index(AUDIT_MONTH)
                if audit_idx > 0:
                    prev_month = all_months_sorted[audit_idx - 1]
                    prev_rent  = rent_by_m.get(prev_month, 0)
                    curr_rent  = rent_by_m.get(AUDIT_MONTH, 0)
                    if prev_rent >= 50 and 0 < curr_rent < prev_rent:
                        decrease = prev_rent - curr_rent
                        pct_drop = decrease / prev_rent
                        if decrease >= 50 and pct_drop >= 0.05:
                            is_ntv = rr_status_lookup.get((prop, unit), "") == "NTV"
                            rule   = "Recurring Rent Decrease (NTV)" if is_ntv else "Recurring Rent Decrease"
                            flags.append(make_flag(prop, unit, resident,
                                rule,
                                f"Recurring rent dropped from ${prev_rent:.0f} ({prev_month}) "
                                f"to ${curr_rent:.0f} ({AUDIT_MONTH}) "
                                f"\u2014 decrease ${decrease:.0f} ({pct_drop*100:.0f}%). "
                                + ("Status NTV \u2014 verify prorated rent is through actual key return date."
                                   if is_ntv else
                                   "Not tied to scheduled move-out \u2014 verify lease/renewal change."),
                                decrease, src))

    # =========================================================================
    # STAGE 2 -- POSTED RENT ROLL AUDIT
    # =========================================================================
    if not df_rent_roll.empty:
        print("  [STAGE 2] Posted Rent Roll Audit ...")
        rr = df_rent_roll.copy()

        # Audit occupied units. 'R' = Renewed (active lease, renewal in progress).
        # Per Daniel: do not skip NTV, eviction, or notice residents.
        rr = rr[rr["Status"].isin(["C", "MTM", "NTV", "R"])]

        # Build move-in lookup from activity, fall back to lease start
        movein_lookup = {}
        if not df_activity.empty:
            for _, row in df_activity.iterrows():
                key = (row["Property"], row["Unit"])
                if key not in movein_lookup and row["Move_In"] is not None:
                    movein_lookup[key] = row["Move_In"]
        if not df_leases.empty:
            for _, row in df_leases.iterrows():
                key = (row["Property"], row["Unit"])
                if key not in movein_lookup:
                    mi = row.get("Lease Start")
                    if mi is not None and pd.notna(mi):
                        movein_lookup[key] = mi

        # 4.1 -- NET RENT INTEGRITY
        for (prop, unit), grp in rr.groupby(["Property", "Unit"]):
            resident = grp["Residents"].iloc[0]
            src      = grp["Source_File"].iloc[0]
            mkt_rent = grp["Market_Rent"].iloc[0]

            # Rent rows: description contains "rent" or "base" but NOT "concession"
            rent_rows = grp[
                grp["Description"].str.lower().str.contains(r"\brent\b|\bbase\b", na=False, regex=True) &
                ~grp["Description"].str.lower().str.contains(r"concession", na=False)
            ]
            net_rent = rent_rows["Amount"].sum()

            if net_rent < 0:
                flags.append(make_flag(prop, unit, resident, "Negative Net Rent",
                    f"Net rent ${net_rent:.2f} -- concession exceeds rent.",
                    abs(net_rent), src))

            elif net_rent == 0:
                move_in = movein_lookup.get((prop, unit))
                today   = pd.Timestamp.today()
                if (move_in is not None and pd.notna(move_in) and
                        (today - pd.Timestamp(move_in)).days <= RECENT_MOVEIN_DAYS):
                    flags.append(make_flag(prop, unit, resident,
                        "$0 Net Rent (Recent Move-in)",
                        f"Net rent $0 â€” moved in {pd.Timestamp(move_in).date()}, "
                        f"within {RECENT_MOVEIN_DAYS} days. "
                        f"Verify first-month timing and that a free-month addendum "
                        f"is on file if applicable.",
                        0.0, src))
                else:
                    flags.append(make_flag(prop, unit, resident,
                        "$0 Net Rent (Not Recent)",
                        f"Net rent is $0 and resident is not a recent move-in. "
                        f"Check whether this is a courtesy officer, employee, or "
                        f"model unit â€” if so, confirm an approved addendum is on file. "
                        f"Otherwise investigate for unauthorized full-offset posting.",
                        0.0, src))

        # 4.2 -- MANUAL CONCESSION (in Transaction List, no projection setup)
        if not df_trans.empty and not df_projection.empty:
            proj_units = set(zip(df_projection["Property"], df_projection["Unit"]))
            active_credits = df_trans[(~df_trans["Is_Reversal"]) & (df_trans["Amount"] > 0)]
            unit_net = active_credits.groupby(["Property", "Unit"])["Amount"].sum()

            for (prop, unit), net in unit_net.items():
                if (prop, unit) not in proj_units and net > 0:
                    sub = active_credits[(active_credits["Property"] == prop) &
                                         (active_credits["Unit"] == unit)]
                    resident = sub["Name"].iloc[0] if not sub.empty else "Unknown"
                    src      = sub["Source_File"].iloc[0] if not sub.empty else ""
                    flags.append(make_flag(prop, unit, resident,
                        "Manual Posting Without Setup",
                        f"${net:.2f} credit in Transaction List but no recurring "
                        f"concession in Projection for this unit.",
                        net, src))

        # 4.3 -- POSTED vs RECURRING MISMATCH
        if not df_projection.empty:
            for (prop, unit), grp in rr.groupby(["Property", "Unit"]):
                # Only count POSITIVE rent rows from the Rent Roll (excludes
                # "Concession - Rent" negative rows which contain "rent" in
                # their description but are not rent charges).  Without this
                # filter, units that have both a gross rent row and a
                # "Concession - Rent" credit row on the Rent Roll produce a
                # net posted_rent that is lower than the gross recurring_rent
                # in the Projection â€” a spurious mismatch.
                rent_rows = grp[
                    grp["Description"].str.lower().str.contains(
                        r"\brent\b|\bbase\b", na=False, regex=True) &
                    (grp["Amount"] > 0)
                ]
                posted_rent = rent_rows["Amount"].sum()

                # Exclude concession/credit rows from the rent query so that
                # recurring_rent reflects the gross rent charge only â€” matching
                # how posted_rent is calculated from the Rent Roll (which already
                # excludes concession rows).  Without this exclusion, units with a
                # "Concession - Rent" entry in the Projection would produce a
                # false "Posted vs Recurring Mismatch" flag every month.
                proj_sub = df_projection[
                    (df_projection["Property"] == prop) &
                    (df_projection["Unit"] == unit) &
                    df_projection["Category"].str.lower().str.contains(
                        r"\brent\b|\bbase\b", na=False, regex=True) &
                    ~df_projection["Category"].str.lower().str.contains(
                        "concession", na=False) &
                    (df_projection["Amount"] > 0)
                ]
                recurring_rent = proj_sub["Amount"].sum()

                if recurring_rent > 0 and posted_rent > 0:
                    # Proration guard â€” bidirectional: skip if EITHER the
                    # projection rent OR the Rent Roll rent is less than 60%
                    # of the other.  This catches two common false-positive
                    # patterns in the audit month column:
                    #   â€¢ New move-in: projection shows a prorated first-month
                    #     amount (small), Rent Roll has full-month gross rent.
                    #   â€¢ Mid-month move-out: Rent Roll shows a prorated final
                    #     charge (small), projection has full-month rent.
                    # Using the SMALLER of the two values vs the LARGER avoids
                    # both directions in a single check.
                    if min(recurring_rent, posted_rent) < max(recurring_rent, posted_rent) * 0.60:
                        # NTV units with large prorations: reclassify as VERIFY rather than skip
                        if grp["Status"].iloc[0] == "NTV" and posted_rent < recurring_rent:
                            resident = grp["Residents"].iloc[0]
                            src      = grp["Source_File"].iloc[0]
                            flags.append(make_flag(prop, unit, resident,
                                "NTV Proration",
                                f"Status NTV. Recurring rent ${recurring_rent:.0f}, "
                                f"posted ${posted_rent:.0f} — likely prorated through scheduled move-out. "
                                f"Verify prorated rent is through actual key return date.",
                                recurring_rent - posted_rent, src))
                        continue
                    var = recurring_rent - posted_rent
                    if abs(var) > 5.0:
                        resident = grp["Residents"].iloc[0]
                        src      = grp["Source_File"].iloc[0]
                        if var > 0:
                            direction_detail = (
                                f"Recurring projection shows rent of ${recurring_rent:.0f}, "
                                f"but the Rent Roll has ${posted_rent:.0f} â€” "
                                f"${var:.0f} higher in the projection. "
                                f"The Rent Roll may reflect a rent reduction not yet "
                                f"updated in the recurring setup, or vice versa."
                            )
                        else:
                            direction_detail = (
                                f"Recurring projection shows rent of ${recurring_rent:.0f}, "
                                f"but the Rent Roll has ${posted_rent:.0f} â€” "
                                f"${abs(var):.0f} lower in the projection. "
                                f"A rent increase may have been applied to the Rent Roll "
                                f"but the recurring charge was not updated to match."
                            )
                        unit_status = grp["Status"].iloc[0]
                        if unit_status == "NTV" and var > 0:
                            flags.append(make_flag(prop, unit, resident,
                                "NTV Proration",
                                f"Status NTV. Recurring rent ${recurring_rent:.0f}, "
                                f"posted ${posted_rent:.0f} — likely prorated through scheduled move-out. "
                                f"Verify prorated rent is through actual key return date.",
                                abs(var), src))
                        else:
                            flags.append(make_flag(prop, unit, resident,
                                "Posted vs Recurring Mismatch",
                                direction_detail,
                                abs(var), src))

        # 4.4 -- MISC TENANT CREDIT REVIEW
        if not df_trans.empty:
            misc_kw = ["misc", "miscellaneous", "adjustment", "write-off",
                       "write off", "reclass", "mccr"]
            misc_mask = (
                df_trans["Description"].str.lower().apply(
                    lambda x: any(k in x for k in misc_kw)
                ) & (~df_trans["Is_Reversal"]) & (df_trans["Amount"] > 0)
            )
            for _, row in df_trans[misc_mask].iterrows():
                flags.append(make_flag(
                    row.get("Property", "?"),
                    row["Unit"],
                    str(row.get("Name", "Unknown")),
                    "Misc Tenant Credit",
                    f"Misc credit ${row['Amount']:.2f} -- '{row['Description']}'. "
                    f"Review individually per specification.",
                    row["Amount"],
                    row.get("Source_File", "")))

    result = pd.DataFrame(flags) if flags else pd.DataFrame()
    print(f"  [REVENUE INTEGRITY ENGINE] Complete -- {len(result)} flags.")
    return result


# ===========================================================================
# SECTION 4B -- FEE SCHEDULE CHECK (Official Fee Sheet Amount Validation)
# ===========================================================================

def run_fee_schedule_check(df_projection: pd.DataFrame) -> pd.DataFrame:
    """
    Compares each unit's existing recurring charges against PROPERTY_FEE_SCHEDULE
    (sourced from the official fee sheet .docx files, March 2026).

    Per Daniel Twito: $1 is the variance cutoff for charge amount comparisons.
    Only flags units where a charge EXISTS but has the WRONG amount.
    Missing charges are handled separately by the 90% Missing Standard Charge rule.

    Skips:
      - Properties without a fee schedule (La Prada)
      - Concession / credit rows (negative amounts)
    """
    print("\n[FEE SCHEDULE CHECK] Validating charge amounts against official fee sheets...")
    flags = []

    if df_projection.empty:
        return pd.DataFrame()

    for prop, prop_grp in df_projection.groupby("Property"):
        schedule = PROPERTY_FEE_SCHEDULE.get(prop)
        if not schedule:
            print(f"  [SKIP] {prop} â€” no fee schedule loaded.")
            continue

        for unit, unit_grp in prop_grp.groupby("Unit"):
            unit_grp = unit_grp.reset_index(drop=True)
            resident = unit_grp["Resident"].iloc[0]
            src      = unit_grp["Source_File"].iloc[0]

            for fee in schedule:
                # Find rows matching this fee's keywords with a positive amount
                matching = unit_grp[
                    unit_grp["Category"].str.lower().apply(
                        lambda cat: any(kw in cat for kw in fee["keywords"])
                    ) & (unit_grp["Amount"] > 0)
                ]

                if matching.empty:
                    # Charge not present â€” Missing Standard Charge handles this
                    continue

                actual_amt = matching["Amount"].iloc[0]
                variance   = abs(actual_amt - fee["amount"])

                # For optional per-item fees (parking, pets, W/D), a resident may
                # have multiple units (e.g. 3 parking spaces at $35 = $105/mo).
                # Skip only if the charge is an exact whole multiple of the per-unit
                # fee rate and the fee is flagged optional in the schedule.
                if (fee.get("optional", False) and fee["amount"] > 0
                        and round(actual_amt % fee["amount"], 2) == 0.0):
                    continue

                # For optional charges (pet rent, parking, W/D), only flag if
                # variance is > $10. Per Daniel: for POT/HP focus on presence/
                # absence, not small amount differences.
                min_variance = 10.0 if fee.get("optional", False) else 1.0
                if variance >= min_variance:
                    flags.append(make_flag(
                        prop, unit, resident,
                        "Fee Schedule Violation",
                        f"'{fee['name']}': Fee schedule = ${fee['amount']:.2f}/mo, "
                        f"Recurring Projection = ${actual_amt:.2f}/mo "
                        f"(variance ${variance:.2f}). Review lease addendum.",
                        variance,
                        src,
                    ))

    result = pd.DataFrame(flags) if flags else pd.DataFrame()
    print(f"  [FEE SCHEDULE CHECK] Complete -- {len(result)} flags.")
    return result


# ===========================================================================
# SECTION 5 -- MANAGER OVERRIDE AUDIT
# ===========================================================================

def run_manager_override_audit(df_edits: pd.DataFrame) -> tuple:
    """
    Uses the Edited Transactions by User report.
    Returns (manager_ranking DataFrame, override_log DataFrame).
    """
    print("\n[OVERRIDE AUDIT] Analysing edited/reversed transactions...")
    if df_edits.empty:
        print("  [SKIP] No edited transactions data.")
        return pd.DataFrame(), pd.DataFrame()

    override_log = df_edits.copy()

    manager_ranking = (
        override_log
        .groupby(["Property", "Manager_Login"])
        .agg(
            Total_Events   =("Event_Type", "count"),
            Reversals      =("Event_Type", lambda x: (x == "Reversal").sum()),
            Amount_Changes =("Event_Type", lambda x: (x == "Amount Change").sum()),
            Total_Impact   =("Revenue_Impact", "sum"),
        )
        .reset_index()
        .sort_values("Total_Impact", ascending=True)
    )

    print(f"  [OVERRIDE] {len(manager_ranking)} managers | "
          f"{len(override_log)} events | "
          f"Revenue impact: ${override_log['Revenue_Impact'].sum():,.2f}")
    return manager_ranking, override_log


# ===========================================================================
# SECTION 5B -- VERIFIED BASELINE SUPPRESSION
# ===========================================================================

def load_verified_baseline() -> pd.DataFrame:
    """Load data/verified_baseline.csv — the accumulated list of flags that
    have been confirmed as 'Verified' (intentional setup) or 'Corrected'
    (fix was applied) by Daniel Twito and property managers.

    Columns: Property, Unit, Rule, Status, Resident, Notes, Month_Verified
    Status values:
      Verified      — intentional setup per approved lease/addendum; suppress.
      Corrected     — fix was applied; keep flag but downgrade + prefix detail.
      Keep_Flagging — known issue, never resolved; always emit at full severity.
    """
    path = os.path.join(DATA_DIR, "verified_baseline.csv")
    if not os.path.exists(path):
        print("  [BASELINE] verified_baseline.csv not found — no suppression applied.")
        return pd.DataFrame()
    try:
        df = pd.read_csv(path, dtype=str).fillna("")
        df["Unit"] = df["Unit"].apply(clean_unit)
        print(f"  [BASELINE] Loaded {len(df)} verified/corrected items.")
        return df
    except Exception as e:
        print(f"  [BASELINE] Could not load verified_baseline.csv: {e}")
        return pd.DataFrame()


def apply_verified_baseline(flags_df: pd.DataFrame,
                             baseline_df: pd.DataFrame) -> pd.DataFrame:
    """Apply verified-baseline suppression to a flags DataFrame.

    Verified     → flag removed from output entirely.
    Corrected    → flag kept but Risk_Level downgraded to MEDIUM and Detail
                   prefixed with '[Previously Corrected <month> — confirm
                   correction still in effect]'.
    Keep_Flagging → no change; flag emitted at full severity.
    """
    if flags_df is None or flags_df.empty:
        return flags_df if flags_df is not None else pd.DataFrame()
    if baseline_df is None or baseline_df.empty:
        return flags_df

    kept = []
    suppressed = 0
    downgraded = 0

    for _, row in flags_df.iterrows():
        match = baseline_df[
            (baseline_df["Property"] == row["Property"]) &
            (baseline_df["Unit"]     == str(row["Unit"])) &
            (baseline_df["Rule"]     == row["Rule"])
        ]
        if match.empty:
            kept.append(row)
            continue

        status = match.iloc[0]["Status"]
        month  = match.iloc[0].get("Month_Verified", "prior month")

        if status == "Verified":
            suppressed += 1
            continue  # suppress completely
        elif status == "Corrected":
            modified = row.copy()
            modified["Detail"]     = (
                f"[Previously Corrected {month} \u2014 confirm correction "
                f"still in effect] " + str(row["Detail"])
            )
            modified["Risk_Level"] = RISK_MEDIUM
            kept.append(modified)
            downgraded += 1
        else:
            kept.append(row)  # Keep_Flagging or unknown

    if suppressed or downgraded:
        print(f"  [BASELINE] Suppressed {suppressed} Verified flags | "
              f"Downgraded {downgraded} Corrected flags to MEDIUM.")

    if not kept:
        return pd.DataFrame(columns=flags_df.columns)
    return pd.DataFrame(kept, columns=flags_df.columns).reset_index(drop=True)


# ===========================================================================
# SECTION 6 -- FINANCIAL EXPOSURE CALCULATOR
# ===========================================================================

def calculate_exposure(flags_df: pd.DataFrame) -> dict:
    """Roll up exception exposure into summary tables."""
    empty = pd.DataFrame()
    if flags_df is None or flags_df.empty:
        return {"by_property": empty, "by_rule": empty,
                "by_risk": empty, "totals": empty}

    df = flags_df.copy()
    df["Amount_Impact"] = pd.to_numeric(df["Amount_Impact"], errors="coerce").fillna(0)

    by_prop = (
        df.groupby(["Property", "Risk_Level"])["Amount_Impact"]
        .agg(Exceptions="count", Total_Exposure="sum")
        .reset_index()
        .sort_values("Total_Exposure", ascending=False)
    )
    by_rule = (
        df.groupby(["Rule", "Risk_Level"])["Amount_Impact"]
        .agg(Count="count", Total_Exposure="sum")
        .reset_index()
        .sort_values("Total_Exposure", ascending=False)
    )
    by_risk = (
        df.groupby("Risk_Level")["Amount_Impact"]
        .agg(Count="count", Total_Exposure="sum")
        .reset_index()
    )
    # Deduped exposure: take the max Amount_Impact per (Property, Unit) to avoid
    # counting the same financial event twice when multiple engines flag the same unit
    # (e.g. John R5 + Daniel "Manual Posting" both flag the same $349 credit).
    # This is a conservative floor; Total_Exposure is the raw ceiling.
    deduped_exposure = round(
        df.groupby(["Property", "Unit"])["Amount_Impact"].max().sum(), 2
    )
    totals = pd.DataFrame([{
        "Total_Units_Audited": df["Unit"].nunique(),
        "Total_Exceptions":    len(df),
        "Total_Exposure":      round(df["Amount_Impact"].sum(), 2),
        "Deduped_Exposure":    deduped_exposure,
        "Critical_Flags":      (df["Risk_Level"] == RISK_CRITICAL).sum(),
        "High_Flags":          (df["Risk_Level"] == RISK_HIGH).sum(),
        "Medium_Flags":        (df["Risk_Level"] == RISK_MEDIUM).sum(),
        "Verify_Flags":        (df["Risk_Level"] == RISK_VERIFY).sum(),
        "Avg_Flags_Per_Unit":   round(len(df) / max(df["Unit"].nunique(), 1), 1),
    }])

    return {"by_property": by_prop, "by_rule": by_rule,
            "by_risk": by_risk, "totals": totals}


# ===========================================================================
# SECTION 7 -- EXCEL EXPORT
# ===========================================================================

# Status options for the resolution workflow dropdown
REVIEW_STATUSES = ["Open", "Reviewed", "Cleared", "Escalated"]

# Row fill colors per status
STATUS_FILLS = {
    "Open":      PatternFill(fill_type=None),                                        # no fill
    "Reviewed":  PatternFill("solid", fgColor="BDD7EE"),                             # light blue
    "Cleared":   PatternFill("solid", fgColor="C6EFCE"),                             # light green
    "Escalated": PatternFill("solid", fgColor="FFEB9C"),                             # light orange
}

# Risk-level header colors
RISK_FILLS = {
    RISK_CRITICAL: PatternFill("solid", fgColor="FF4B4B"),
    RISK_HIGH:     PatternFill("solid", fgColor="FFA500"),
    RISK_MEDIUM:   PatternFill("solid", fgColor="FFD700"),
    RISK_VERIFY:   PatternFill("solid", fgColor="A8D8A8"),   # light green — verify only
}

HEADER_FILL  = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER  = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)


def _add_review_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Prepend Status='Open' and Notes='' columns to a flags DataFrame."""
    if df.empty:
        return df
    out = df.copy()
    out.insert(0, "Notes",  "")
    out.insert(0, "Status", "Open")
    return out


def _format_flag_sheet(ws, status_col_idx: int) -> None:
    """
    Apply formatting to a flag worksheet:
      - Dark blue header row with white bold text
      - Freeze header row
      - Dropdown validation on Status column
      - Row color based on current Status value
      - Auto-fit column widths (capped at 60)
      - Bottom border on every data row
    """
    max_row = ws.max_row
    max_col = ws.max_column

    # --- Header row ---
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # --- Dropdown validation on Status column ---
    status_letter = get_column_letter(status_col_idx)
    dv = DataValidation(
        type="list",
        formula1=f'"{",".join(REVIEW_STATUSES)}"',
        allow_blank=False,
        showDropDown=False,  # False = show the arrow in Excel
    )
    dv.sqref = f"{status_letter}2:{status_letter}{max(max_row, 2)}"
    ws.add_data_validation(dv)

    # --- Row color + border based on Status value ---
    for row_idx in range(2, max_row + 1):
        status_val = ws.cell(row=row_idx, column=status_col_idx).value or "Open"
        fill = STATUS_FILLS.get(status_val, STATUS_FILLS["Open"])
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill   = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    # --- Auto-fit column widths ---
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        header_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
        max_len    = header_len
        for row_idx in range(2, min(max_row + 1, 200)):   # sample first 200 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 60))
        ws.column_dimensions[col_letter].width = max(max_len + 2, 10)


def export_to_excel(johns_flags, daniels_flags, fee_flags,
                    manager_ranking, override_log, exposure,
                    output_dir=None, audit_month="") -> bytes:
    month_slug = audit_month.replace(" ", "") if audit_month else ""
    ts_suffix  = datetime.now().strftime("%Y%m%d_%H%M")
    ts = f"{month_slug}_{ts_suffix}" if month_slug else ts_suffix
    if output_dir is not None:
        os.makedirs(output_dir, exist_ok=True)
        out_path = os.path.join(output_dir, f"LNJ_Audit_{ts}.xlsx")
        write_target = out_path
    else:
        write_target = io.BytesIO()

    # Add Status/Notes resolution columns to all flag sheets
    j_flags   = _add_review_columns(johns_flags)
    d_flags   = _add_review_columns(daniels_flags)
    f_flags   = _add_review_columns(fee_flags)

    all_flags_raw = pd.concat([johns_flags, daniels_flags, fee_flags], ignore_index=True) \
                    if not (johns_flags.empty and daniels_flags.empty and fee_flags.empty) \
                    else pd.DataFrame()
    a_flags = _add_review_columns(all_flags_raw)

    risk_order = {RISK_CRITICAL: 0, RISK_HIGH: 1, RISK_MEDIUM: 2}

    # --- Track which sheets need flag formatting and where Status col is ---
    flag_sheets = {}   # sheet_name -> status_col_index (1-based)

    with pd.ExcelWriter(write_target, engine="openpyxl") as writer:
        if not exposure.get("totals", pd.DataFrame()).empty:
            exposure["totals"].to_excel(writer, sheet_name="Executive Summary", index=False)

        if not a_flags.empty:
            sorted_flags = a_flags.copy()
            sorted_flags["_rank"] = sorted_flags["Risk_Level"].map(risk_order).fillna(3)
            final = (sorted_flags
                     .sort_values(["_rank", "Amount_Impact"], ascending=[True, False])
                     .drop(columns=["_rank"]))
            final.to_excel(writer, sheet_name="All Exceptions", index=False)
            flag_sheets["All Exceptions"] = list(final.columns).index("Status") + 1

        if not j_flags.empty:
            j_flags.to_excel(writer, sheet_name="Concession Audit Engine", index=False)
            flag_sheets["Concession Audit Engine"] = list(j_flags.columns).index("Status") + 1

        if not d_flags.empty:
            d_flags.to_excel(writer, sheet_name="Revenue Integrity Engine", index=False)
            flag_sheets["Revenue Integrity Engine"] = list(d_flags.columns).index("Status") + 1

        if not f_flags.empty:
            f_flags.to_excel(writer, sheet_name="Fee Schedule Violations", index=False)
            flag_sheets["Fee Schedule Violations"] = list(f_flags.columns).index("Status") + 1

        if not exposure.get("by_property", pd.DataFrame()).empty:
            exposure["by_property"].to_excel(writer, sheet_name="Exposure by Property", index=False)
        if not exposure.get("by_rule", pd.DataFrame()).empty:
            exposure["by_rule"].to_excel(writer, sheet_name="Exposure by Rule", index=False)

        if not manager_ranking.empty:
            manager_ranking.to_excel(writer, sheet_name="Manager Ranking", index=False)
        if not override_log.empty:
            override_log.to_excel(writer, sheet_name="Override Detail Log", index=False)

        # openpyxl requires at least one visible sheet
        if not writer.sheets:
            pd.DataFrame({"Info": ["No audit data — upload files and run again."]}).to_excel(
                writer, sheet_name="No Data", index=False
            )

    # --- Post-process with openpyxl for formatting ---
    from openpyxl import load_workbook
    if output_dir is not None:
        wb = load_workbook(out_path)
    else:
        write_target.seek(0)
        wb = load_workbook(write_target)

    for sheet_name, status_col in flag_sheets.items():
        if sheet_name in wb.sheetnames:
            _format_flag_sheet(wb[sheet_name], status_col)

    # Light formatting on non-flag sheets (header only)
    for sheet_name in wb.sheetnames:
        if sheet_name not in flag_sheets:
            ws = wb[sheet_name]
            for cell in ws[1]:
                cell.fill      = HEADER_FILL
                cell.font      = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            ws.freeze_panes = "A2"
            for col_idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
                for row_idx in range(2, min(ws.max_row + 1, 100)):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val:
                        max_len = max(max_len, min(len(str(val)), 50))
                ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

    if output_dir is not None:
        wb.save(out_path)
        print(f"\n[EXPORT] Saved -> {out_path}")
        with open(out_path, "rb") as f:
            return f.read()
    else:
        out_buf = io.BytesIO()
        wb.save(out_buf)
        print(f"\n[EXPORT] Excel generated in memory ({ts}).")
        return out_buf.getvalue()


# ===========================================================================
# SECTION 8 -- MAIN ORCHESTRATOR
# ===========================================================================

def run_full_audit(uploaded_files=None, audit_month=None) -> dict:
    """
    Single entry point for both CLI and Streamlit app.
    uploaded_files: dict of {key: [UploadedFile, ...]} for Streamlit upload mode.
                   When None, reads from data/ subfolders (CLI mode).
    audit_month:   Override AUDIT_MONTH for the projection loader (e.g. "Jul 2026").
    Returns a dict of all result DataFrames, plus "excel_bytes" (bytes).
    """
    _month    = audit_month.strip() if audit_month else AUDIT_MONTH
    is_upload = uploaded_files is not None
    uf        = uploaded_files or {}

    # In upload mode: missing keys → empty list (no data for that type).
    # In CLI mode: pass None → loaders use disk folders.
    def _uf(key):
        return uf.get(key) or [] if is_upload else None

    print("=" * 60)
    print("  LiveNjoy Residential -- ResMan Audit Bot")
    print(f"  Audit Month : {_month}")
    print(f"  Run Time    : {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print("=" * 60)

    # -- Ingest ---------------------------------------------------------------
    mode_label = "uploaded files" if is_upload else "data/ subfolders"
    print(f"\n[INGEST] Loading from {mode_label} ...")
    df_trans      = load_transaction_list(DIRS["transactions"],         uploaded_files=_uf("transactions"))
    df_leases     = load_leases(DIRS["leases"],                         uploaded_files=_uf("leases"))
    df_edits      = load_edits(DIRS["edits"],                           uploaded_files=_uf("edits"))
    df_projection = load_transaction_projection(DIRS["recurring"],      audit_month=_month, uploaded_files=_uf("recurring"))
    df_proj_full  = load_transaction_projection_all_months(DIRS["recurring"],               uploaded_files=_uf("recurring"))
    df_rent_roll  = load_rent_roll(DIRS["rent_rolls"],                  uploaded_files=_uf("rent_rolls"))
    df_activity   = load_resident_activity(DIRS["activity"],            uploaded_files=_uf("activity"))
    mkt_rent_lookup = load_market_rent_schedule(DIRS["market_rent_schedule"], uploaded_files=_uf("market_rent_schedule"))

    print(f"\n  Loaded  : {len(df_trans)} transaction rows | "
          f"{len(df_leases)} lease rows | "
          f"{len(df_projection)} projection rows | "
          f"{len(df_rent_roll)} rent roll charge rows | "
          f"{len(mkt_rent_lookup)} market-rent-schedule units")

    # -- Engines --------------------------------------------------------------
    concession_flags   = run_concession_audit_engine(df_trans, df_leases,
                                     df_projection=df_projection,
                                     df_rent_roll=df_rent_roll)
    revenue_integrity_flags = run_revenue_integrity_engine(
        df_projection, df_rent_roll, df_trans, df_leases, df_activity, df_proj_full,
        market_rent_lookup=mkt_rent_lookup
    )
    fee_flags     = run_fee_schedule_check(df_projection)
    manager_ranking, override_log = run_manager_override_audit(df_edits)

    # -- Apply verified baseline suppression ----------------------------------
    _baseline               = load_verified_baseline()
    concession_flags        = apply_verified_baseline(concession_flags, _baseline)
    revenue_integrity_flags = apply_verified_baseline(revenue_integrity_flags, _baseline)
    fee_flags               = apply_verified_baseline(fee_flags, _baseline)

    # -- Combine + exposure ---------------------------------------------------
    all_flags = (pd.concat([concession_flags, revenue_integrity_flags, fee_flags], ignore_index=True)
                 if not (concession_flags.empty and revenue_integrity_flags.empty and fee_flags.empty)
                 else pd.DataFrame())
    exposure  = calculate_exposure(all_flags)

    if not all_flags.empty:
        t = exposure["totals"].iloc[0]
        print(f"\n[RESULTS] {int(t['Total_Units_Audited'])} units | "
              f"{int(t['Total_Exceptions'])} exceptions | "
              f"${t['Total_Exposure']:,.2f} exposure | "
              f"{int(t['Critical_Flags'])} CRITICAL")

    # -- Export ---------------------------------------------------------------
    # Upload mode: generate Excel in memory only (no disk write).
    # CLI mode: write to output/ folder on disk.
    _out_dir    = None if is_upload else OUTPUT_DIR
    excel_bytes = export_to_excel(concession_flags, revenue_integrity_flags, fee_flags,
                                  manager_ranking, override_log, exposure,
                                  output_dir=_out_dir, audit_month=_month)

    print("\n[DONE] Audit complete.")
    return {
        "df_trans":        df_trans,
        "df_leases":       df_leases,
        "df_projection":   df_projection,
        "df_rent_roll":    df_rent_roll,
        "df_activity":     df_activity,
        "df_edits":        df_edits,
        "concession_flags":        concession_flags,
        "fee_flags":               fee_flags,
        "revenue_integrity_flags": revenue_integrity_flags,
        "all_flags":       all_flags,
        "manager_ranking": manager_ranking,
        "override_log":    override_log,
        "exposure":        exposure,
        "excel_bytes":     excel_bytes,
    }


# -- CLI ----------------------------------------------------------------------
if __name__ == "__main__":
    run_full_audit()
